from loguru import logger
from tta_helpers import with_gc, setup_logging, load_data
from multiprocessing import Queue, Event
import os, sys
from typing import Tuple
from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter
import platform
import subprocess
from optimizer_result_model import OptimizerResult
import pickle


@with_gc
def run_analysis_threaded(
    files_list=[],
    strategy_settings={},
    open_files=False,
    results_queue: Queue = None,
    cancel_flag: Event = None,  # type: ignore
    create_excel=True,
    news_events={},
    optimizer_result: OptimizerResult = None,
):
    global logger
    logger = setup_logging(logger, "DEBUG")
    try:
        # initialize df_dicts
        df_dicts = {}

        for file in files_list:
            strategy = (
                "-SINGLE_MODE-"
                if "-SINGLE_MODE-" in strategy_settings
                else os.path.basename(file)
            )
            settings = strategy_settings[strategy]
            logger.info(f"Creating heatmap for {file}")
            result_dicts = create_excel_file(
                file,
                settings,
                open_files,
                create_excel,
                news_events,
                cancel_flag,
                results_queue,
            )

            # check for cancel flag to stop thread
            if cancel_flag is not None and cancel_flag.is_set():
                cancel_flag.clear()
                if results_queue:
                    results_queue.put(("-BACKTEST_CANCELED-", "-RUN_ANALYSIS-"))
                return

            source = os.path.splitext(os.path.basename(file))[0]
            for right_type, day_dict in result_dicts.items():
                if right_type not in df_dicts:
                    df_dicts[right_type] = {
                        "All": {},
                        "Mon": {},
                        "Tue": {},
                        "Wed": {},
                        "Thu": {},
                        "Fri": {},
                    }
                for day, df_dict in day_dict.items():
                    df_dicts[right_type][day][source] = df_dict

        for _best in ["Best P/C", "Best P/C Gap Up", "Best P/C Gap Down"]:
            df_dicts[_best] = {
                "All": {},
                "Mon": {},
                "Tue": {},
                "Wed": {},
                "Thu": {},
                "Fri": {},
            }

            # combine the put and call dfs into 1 dict for determining the best time
            # from among both individual datasets
            for _right in ["Puts", "Calls"]:
                if _best.endswith("Gap Up"):
                    _right += " Gap Up"
                elif _best.endswith("Gap Down"):
                    _right += " Gap Down"

                if _right in df_dicts:
                    for _day, _day_dict in df_dicts[_right].items():
                        for _source, _df_dict in _day_dict.items():
                            df_dicts[_best][_day][
                                f"{_right.removesuffix("s")}||{_source}"
                            ] = _df_dict

        logger.info("Analysis Finished")
        if results_queue:
            logger.debug("Pickling results...")
            # Get the directory of the current script
            if getattr(sys, "frozen", False):
                # The application is running in a bundle (PyInstaller)
                current_dir = os.path.dirname(sys.executable)
            else:
                # The application is running in a normal Python environment
                current_dir = os.path.dirname(os.path.abspath(__file__))
            # Create a 'data' directory if it doesn't exist
            data_dir = os.path.join(current_dir, "data")
            os.makedirs(data_dir, exist_ok=True)
            results_path = os.path.join(data_dir, "results.pkl")
            with open(results_path, "wb") as f:
                pickle.dump(df_dicts, f)
            logger.debug("Results pickled")
            results_queue.put(("-RUN_ANALYSIS_END-", ("-RESULTS_PATH-", results_path)))

        if optimizer_result is not None:
            optimizer_result.run_analysis_result = df_dicts
            return optimizer_result
        return df_dicts
    except Exception as e:
        logger.exception("Error in run_analysis_threaded")
        if results_queue:
            results_queue.put(("-RUN_ANALYSIS_END-", e))
        else:
            return e


def create_excel_file(
    file,
    settings,
    open_files,
    create_excel=True,
    news_events={},
    cancel_flag: Event = None,  # type: ignore
    results_queue: Queue = None,
) -> dict:
    try:
        calc_type = settings["-CALC_TYPE-"]
        short_avg_period = settings["-AVG_PERIOD_1-"]
        short_weight = settings["-PERIOD_1_WEIGHT-"] / 100
        long_avg_period = settings["-AVG_PERIOD_2-"]
        long_weight = settings["-PERIOD_2_WEIGHT-"] / 100
        top_x = settings["-TOP_X-"]
        weekday_exclusions = []
        news_date_exclusions = []
        if settings["-APPLY_EXCLUSIONS-"] != "Walk Forward Test":
            weekday_exclusions = settings["-WEEKDAY_EXCLUSIONS-"]
            # get list of news event dates to skip.
            for release, date_list in news_events.items():
                if release in settings["-NEWS_EXCLUSIONS-"]:
                    news_date_exclusions += date_list

        # load the data
        result = load_data(file)
        if result:
            df, start_date, end_date = result
            filtered_df = df[
                (~df["Day of Week"].isin(weekday_exclusions))
                & (~df["EntryTime"].dt.date.isin(news_date_exclusions))
            ]
        else:
            return

        # path and original filename
        path = os.path.join(os.path.dirname(file), "data", "heatmaps")
        org_filename = os.path.splitext(os.path.basename(file))[0]
        os.makedirs(path, exist_ok=True)

        # Create filename
        filename = os.path.join(
            path,
            (
                f"{org_filename}-TWAvg({calc_type})_{short_avg_period}mo({short_weight * 100:.0f})-{long_avg_period}mo({long_weight * 100:.0f})_{start_date} -"
                f" {end_date}.xlsx"
            ),
        )
        filename = filename if create_excel else BytesIO()  # don't make the file
        # Create a Pandas Excel writer using XlsxWriter as the engine
        with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:

            # Get the xlsxwriter workbook
            workbook = writer.book

            # get the sheets for day of week
            day_to_num = {
                "Monday": 1,
                "Tuesday": 2,
                "Wednesday": 3,
                "Thursday": 4,
                "Friday": 5,
                "Saturday": 6,
                "Sunday": 7,
            }

            days_sorted = ["All"]
            if settings["-IDV_WEEKDAY-"]:
                # This gets the unique days of the week from the DataFrame, then sorts them based on the numerical value
                days_sorted = days_sorted + sorted(
                    [
                        d
                        for d in filtered_df["Day of Week"].unique()
                        if d not in settings["-WEEKDAY_EXCLUSIONS-"]
                    ],
                    key=lambda day: day_to_num[day],
                )

            df_dicts = {"Put-Call Comb": {}}
            if settings["-PUT_OR_CALL-"]:
                df_dicts["Puts"] = {}
                df_dicts["Calls"] = {}

            if settings["-GAP_ANALYSIS-"]:
                for strat in df_dicts.copy():
                    df_dicts[f"{strat} Gap Up"] = {}
                    df_dicts[f"{strat} Gap Down"] = {}

            gap_error = False
            for strat in df_dicts.copy():
                for day in days_sorted:
                    # check for cancel flag to stop thread
                    if cancel_flag is not None and cancel_flag.is_set():
                        return

                    # filter for the weekday
                    # we will keep a dataset with exclusions filtered out and one
                    # with all the data.  We will sort and filter those both for
                    # the analysis but the analysis will only happen on the filtered df
                    # this will allow us to store either the filtered df that has the
                    # exclusions removed or the original df that was filtered for the analysis
                    # type, but still has the excluded events.  This filtered, but non-excluded
                    # df will be what is used for the WF test.  This allows events/weekday exclusions
                    # to be done for analysis only, but still traded during the WF test.
                    if day == "All":
                        _df = df
                        _filtered_df = filtered_df
                    else:
                        _df = df[df["Day of Week"] == day]
                        _filtered_df = filtered_df[filtered_df["Day of Week"] == day]

                    # filter for calls/puts
                    if strat.startswith("Puts"):
                        _df = _df[_df["OptionType"] == "P"]
                        _filtered_df = _filtered_df[_filtered_df["OptionType"] == "P"]
                    elif strat.startswith("Calls"):
                        _df = _df[_df["OptionType"] == "C"]
                        _filtered_df = _filtered_df[_filtered_df["OptionType"] == "C"]

                    # filter for gaps
                    _gap_type = "Gap%" if settings["-GAP_TYPE-"] == "%" else "Gap"
                    try:
                        if strat.endswith("Gap Up"):
                            _df = _df[_df[_gap_type] > settings["-GAP_THRESHOLD-"]]
                            _filtered_df = _filtered_df[
                                _filtered_df[_gap_type] > settings["-GAP_THRESHOLD-"]
                            ]
                        elif strat.endswith("Gap Down"):
                            _df = _df[_df[_gap_type] < -settings["-GAP_THRESHOLD-"]]
                            _filtered_df = _filtered_df[
                                _filtered_df[_gap_type] < -settings["-GAP_THRESHOLD-"]
                            ]
                    except KeyError:
                        # gap data did not load, maybe no internet
                        _df = pd.DataFrame(columns=df.columns)
                        _filtered_df = pd.DataFrame(columns=df.columns)
                        if not gap_error:
                            gap_error = True  # only notify once
                            if results_queue:
                                results_queue.put(
                                    (
                                        "-ERROR-",
                                        "Gap data could not be loaded!\nAnalysis will continue without it.",
                                    )
                                )

                    # run the analysis
                    if settings["-APPLY_EXCLUSIONS-"] != "Walk Forward Test":
                        # exclusions are only for WF, so use the non-filtered df
                        df_output, df_output_1mo_avg = analyze(
                            _df, settings, news_events
                        )
                        # store the results and the original df in case we need it later
                        df_dicts[strat][day[:3]] = {
                            "org_df": _df,
                            "result_df": df_output,
                        }
                    else:
                        # otherwise we use the filtered/excluded df for analysis
                        df_output, df_output_1mo_avg = analyze(
                            _filtered_df, settings, news_events
                        )
                        # store the results and the original df in case we need it later
                        if settings["-APPLY_EXCLUSIONS-"] == "Analysis":
                            # since we are only excluded from analysis we will store the non-filtered df
                            df_dicts[strat][day[:3]] = {
                                "org_df": _df,
                                "result_df": df_output,
                            }
                        else:
                            # otherwise we are excluding from both so we can should store the filtered df
                            df_dicts[strat][day[:3]] = {
                                "org_df": _filtered_df,
                                "result_df": df_output,
                            }

                    # create the sheets
                    if not settings["-PASSTHROUGH_MODE-"]:
                        df_output.to_excel(
                            writer, sheet_name=f"{strat}_{day[:3]}", index=False
                        )
                        df_output_1mo_avg.to_excel(
                            writer, sheet_name=f"{strat}_1mo-{day[:3]}", index=False
                        )

            # use All df from Put/Call Combined for row and col lengths
            df_output = df_dicts["Put-Call Comb"]["All"]["result_df"]
            # Set the PCR columns to percentage format
            percent_format = workbook.add_format(
                {"num_format": "0.00%", "align": "center"}
            )
            top_x_format = workbook.add_format(
                {"bold": 1, "font_color": "#FFFFFF"}
            )  # white
            for row in range(
                2, len(df_output) + 2
            ):  # +2 because Excel's index starts from 1 and there is a header row
                for worksheet in writer.sheets.values():
                    # Apply a conditional format to the PCR cells in the current row
                    worksheet.conditional_format(
                        f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                        {
                            "type": "3_color_scale",
                            "min_color": "red",
                            "mid_color": "yellow",
                            "max_color": "green",
                        },
                    )
                    # Format top x values in bold white text
                    if top_x > 0:
                        worksheet.conditional_format(
                            f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                            {
                                "type": "top",
                                "value": top_x,
                                "format": top_x_format,
                            },
                        )
                    if calc_type == "PCR":
                        worksheet.set_row(row - 1, None, percent_format)

            # Adjust the column widths
            for column in df_output:
                column_length = max(
                    df_output[column].astype(str).map(len).max() + 1, len(column) + 1
                )
                col_idx = df_output.columns.get_loc(column)
                for worksheet in writer.sheets.values():
                    worksheet.set_column(col_idx, col_idx, column_length)

        # open file in excel
        if open_files:
            try:
                if platform.system() == "Windows":
                    os.startfile(filename)
                elif (
                    platform.system() == "Darwin"
                ):  # This is the value returned for macOS
                    subprocess.Popen(["open", filename])
                else:
                    subprocess.call(("xdg-open", filename))  # linux
            except:
                pass

        return df_dicts
    except Exception as e:
        logger.exception("Error in create_excel_file")
        raise e


def analyze(
    df: pd.DataFrame,
    settings: dict,
    news_events: dict = {},
) -> Tuple[
    pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame
]:
    try:
        if df.empty or settings["-PASSTHROUGH_MODE-"]:
            return pd.DataFrame(columns=["Date Range"]), pd.DataFrame(
                columns=["Date Range"]
            )
        short_avg_period = settings["-AVG_PERIOD_1-"]
        long_avg_period = settings["-AVG_PERIOD_2-"]
        short_weight = settings["-PERIOD_1_WEIGHT-"] / 100
        long_weight = settings["-PERIOD_2_WEIGHT-"] / 100
        calc_type = settings["-CALC_TYPE-"]
        agg_type = "".join(word[0] for word in settings["-AGG_TYPE-"].split("-"))

        def calculate_avg_pnl(df: pd.DataFrame) -> float:
            if df.columns[0] == "Date Opened":  # OO BT data
                return df["P/L"].sum() / df["No. of Contracts"].sum()
            elif df.columns[0] == "TradeID":  # BYOB BT data
                df["P/L"] = (
                    df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
                ) * 100
                return df["P/L"].mean()
            else:
                raise ValueError("Unknown dataset type")

        def calculate_pcr(df: pd.DataFrame) -> float:
            if df.columns[0] == "Date Opened":  # OO BT data
                return df["P/L"].sum() / (df["Premium"] * df["No. of Contracts"]).sum()
            elif df.columns[0] == "TradeID":  # BYOB BT data
                df["P/L"] = df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
                return df["P/L"].sum() / df["Premium"].sum()
            else:
                raise ValueError("Unknown dataset type")

        def calculate_rolling_averages(df, short_avg_period, long_avg_period, agg_type):
            if agg_type == "W":
                short_avg_period = int(short_avg_period * 4.33)
                long_avg_period = int(long_avg_period * 4.33)
            elif agg_type == "SM":  # semi-monthly
                short_avg_period = int(short_avg_period * 2)
                long_avg_period = int(long_avg_period * 2)

            short_avg = df.rolling(short_avg_period, min_periods=1).mean()
            long_avg = df.rolling(long_avg_period, min_periods=1).mean()
            weighted_avg = short_weight * short_avg + long_weight * long_avg
            return weighted_avg

        def create_output_labels(df, long_avg_period, start_date, end_date, agg_type):
            output_labels = pd.DataFrame(index=df.index)
            for i, (date, row) in enumerate(df.iterrows()):
                if agg_type == "M":
                    current_period_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
                    previous_period_start = (
                        current_period_end - pd.DateOffset(months=long_avg_period - 1)
                    ).replace(day=1)
                elif agg_type == "W":
                    current_period_end = date.to_timestamp() + pd.offsets.Week(
                        weekday=6
                    )
                    previous_period_start = current_period_end - pd.DateOffset(
                        weeks=int(long_avg_period * 4.33)
                    )
                elif agg_type == "SM":  # Semi-Monthly
                    if date.day <= 15:
                        current_period_end = pd.Timestamp(date.year, date.month, 15)
                        previous_period_start = current_period_end - pd.DateOffset(
                            months=long_avg_period
                        )
                    else:
                        current_period_end = pd.Timestamp(
                            date.year, date.month, date.days_in_month
                        )
                        previous_period_start = (
                            current_period_end
                            - pd.DateOffset(months=long_avg_period - 1)
                        ).replace(day=1)
                    # if previous_period_start.day > 15:
                    #     previous_period_start = previous_period_start.replace(day=16)
                    # else:
                    #     previous_period_start = previous_period_start.replace(day=1)
                else:
                    current_period_end = date.to_timestamp() + pd.offsets.DateOffset(
                        freq=agg_type
                    )
                    previous_period_start = current_period_end - pd.DateOffset(
                        freq=agg_type, periods=long_avg_period - 1
                    )

                if i == 0:
                    date_range_label = f"{end_date} - {previous_period_start.date()}"
                elif i == len(df) - 1:
                    date_range_label = f"{current_period_end.date()} - {start_date}"
                else:
                    date_range_label = (
                        f"{current_period_end.date()} - {previous_period_start.date()}"
                    )
                output_labels.loc[date, "Date Range"] = date_range_label
            return output_labels

        def perform_analysis(df_grouped):
            if calc_type == "PCR":
                df_calc = df_grouped.apply(calculate_pcr, include_groups=False)
            elif calc_type == "PnL":
                df_calc = df_grouped.apply(calculate_avg_pnl, include_groups=False)
            else:
                raise ValueError("Invalid calc_type. Expected 'PCR' or 'PnL'.")

            if isinstance(df_calc.index, pd.MultiIndex):
                df_calc = df_calc.unstack(level=-1)

            weighted_avg = calculate_rolling_averages(
                df_calc, short_avg_period, long_avg_period, agg_type
            )
            one_month_avg = df_calc.rolling(
                1 if agg_type == "M" else 2 if agg_type == "SM" else 4, min_periods=1
            ).mean()

            weighted_avg.sort_index(ascending=False, inplace=True)
            one_month_avg.sort_index(ascending=False, inplace=True)

            if isinstance(weighted_avg, pd.Series):
                weighted_avg = weighted_avg.to_frame()

            if calc_type == "PCR":
                weighted_avg = weighted_avg.apply(lambda x: round(x, 4))
                one_month_avg = one_month_avg.apply(lambda x: round(x, 4))
            elif calc_type == "PnL":
                weighted_avg = weighted_avg.apply(lambda x: round(x, 2))
                one_month_avg = one_month_avg.apply(lambda x: round(x, 2))

            output_labels = create_output_labels(
                weighted_avg, long_avg_period, start_date, end_date, agg_type
            )
            one_month_avg_labels = create_output_labels(
                one_month_avg, 1, start_date, end_date, agg_type
            )

            df_output = pd.concat([output_labels, weighted_avg], axis=1)
            df_output_1mo_avg = pd.concat([one_month_avg_labels, one_month_avg], axis=1)

            return df_output, df_output_1mo_avg

        # get list of news event dates to skip.
        news_date_exclusions = []
        if settings["-APPLY_EXCLUSIONS-"] != "Walk Forward Test":
            for release, date_list in news_events.items():
                if release in settings["-NEWS_EXCLUSIONS-"]:
                    news_date_exclusions += date_list

            # filter df for news exclusions
            df = df[~df["EntryTime"].dt.date.isin(news_date_exclusions)]

            # filter for weekday exclusions
            df = df[~df["Day of Week"].isin(settings["-WEEKDAY_EXCLUSIONS-"])]

        if agg_type == "SM":
            # Custom function to create semi-monthly periods
            def semi_monthly_period(date):
                return pd.Timestamp(
                    date.year, date.month, 15 if date.day <= 15 else date.days_in_month
                )

            df["period"] = df["EntryTime"].apply(semi_monthly_period)
            df_grouped_combined = df.groupby(["period", "Time"])
        else:
            df_grouped_combined = df.groupby(
                [df["EntryTime"].dt.to_period(agg_type), "Time"]
            )
        start_date = df["EntryTime"].min().date()
        end_date = df["EntryTime"].max().date()

        if df.empty:
            df_output_combined, df_output_1mo_avg_combined = pd.DataFrame(
                columns=["Date Range"]
            ), pd.DataFrame(columns=["Date Range"])
        else:
            df_output_combined, df_output_1mo_avg_combined = perform_analysis(
                df_grouped_combined
            )

        return (
            df_output_combined,
            df_output_1mo_avg_combined,
        )
    except Exception as e:
        logger.exception("Error in analyze")
        raise e
