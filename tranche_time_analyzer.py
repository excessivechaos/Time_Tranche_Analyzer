import pandas as pd
import numpy as np
import datetime as dt
from openpyxl.utils import get_column_letter
import PySimpleGUI as sg
import os, threading, queue, gc, functools
import subprocess, platform
from typing import Tuple
import ctypes
import base64
from io import BytesIO
from PIL import Image, ImageTk
from dateutil import parser
from dateutil.relativedelta import relativedelta
import matplotlib
import uuid
import csv

matplotlib.use("TkAgg")
import matplotlib.pyplot as plt


# make app dpi aware
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

icon = b"iVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAYAAABccqhmAAAQIElEQVR4Xu2dX2hcdRbHz81MpkmbSZtp122aamttYesW/yCL4CJ2QbesLyJ92CdBkcr6B0EfFnGh4oOUgu3LIoIgiyI+6SI+lIKFRVYLXSnIUrRaLWxqV23axDQ17WRmcpd01zgz2Zj2l3PP72TuJ6/NPd/z+5yTD7fk5k6SpmkqfEEAArkkkCCAXM6dQ0PgMgEEwCJAIMcEEECOh8/RIYAA2AEI5JgAAsjx8Dk6BBAAOwCBHBNAADkePkeHAAJgByCQYwIIIMfD5+gQQADsAARyTAAB5Hj4HB0CCIAdgECOCSCAHA+fo0MAAbADEMgxAQSQ4+FzdAggAHYAAjkmgAByPHyODgEEwA5AIMcEEECOh8/RIYAA2AEI5JgAAsjx8Dk6BBAAOwCBHBNAADkePkeHAAJgByCQYwIIIMfD5+gQQADsAARyTAAB5Hj4HB0CwQL428gRuTQ9BUEIQCAiget6B+WX/ZuDOwgWwO5P/yzjtYngYC6EAAQWT+COyq3y+/W/Cy6EAILRcSEE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYgmgD+8q+/yoX6ZLSDEwwBCIhs698iv/nZ7cEogl8JFpzIhRCAgBsCCMDNKGgEAvYEEIA9cxIh4IYAAnAzChqBgD0BBGDPnEQIuCGAANyMgkYgYE8AAdgzJxECbgggADejoBEI2BNAAPbMSYSAGwLBAnjq9CEZa1xycxCNRv7081/LlmUDc0qlskdEGhoRbmok8qiIrJ7tp/rhITm/9xk3/Wk0Utxyowzse72l1InqqLzw7WGN8m5qVAo9sn/o7qB+EEATNgSAAIJ+iiJfhACUBoAAEIDSKpmWQQBKuBEAAlBaJdMyCEAJNwJAAEqrZFoGASjhRgAIQGmVTMsgACXcCAABKK2SaRkEoIQbASAApVUyLYMAlHAjAASgtEqmZRCAEm4EgACUVsm0TBQBvHT2qEw0pkwPmnXYA5VtMtRdnhOTypsd+CTgfSLSP3vW2rGj8v2br2SN2LR+Yf0GKT/2bEvmV7UJeWP0mGkfWYeVCyV5fM1tQTHBTwIGpXERBCDgigACcDUOmoGALQEEYMubNAi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEggWwBfVManLtG23GadtKK2U3qQ4J+Wz6qikkmacblt+U2mVlJJCU+jMx7yN2DaReVpJRAbbUqoi8k3mybYBM3NcHxQZLIA8vRFo1/ABqXWY7Pas2y6Dxb6mpflUUnk7aIn8XjQoiTzc1t4pSeU1vy0HddYviTwZdCUCaMI235OACCBotxxchAAWGgICQAD/I8AdwEI/LH7/nTsAldlwB8B/AVQWybwIAlBBjgAQgMoimRdBACrIEQACUFkk8yIIQAU5AkAAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5EQSgghwBIACVRTIvEkEABydOysXpuvlRswy8q+9aqRR650S8O35CGh32KPA95Y3S1zXzqOwPX2cllU+yxGteO5GZtzvd2pZ7XlL52LyXLAMTWSYitwdFBD8IFJTGRRCAgCsCCMDVOGgGArYEEIAtb9Ig4IoAAnA1DpqBgC0BBGDLmzQIuCKAAFyNg2YgYEsgWADnHrpXps+dse0247RVe1+V7q03z0nJw/sAPpr8WmY+7KWTvq4vrZTn1t7ZdiReCNIMBAE00UAACGBpCjDCg0DcASzNVfmh6/ZXgnEHsJTniQBUpscdAHcAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5EQSgghwBIACVRTIvggBUkCMABKCySOZFEIAKcgSAAFQWybwIAlBBjgAQgMoimReJIIDze5+R6fEx86NmGVj+wx+lcN0NcyJePHNE6mlnfQ7iI2tuaXn5yfHqOXnnu8+zxGtee233CnmwclNb7oikctC8l2wDl0siO4Migp8EDErjIghAwBUBBOBqHDQDAVsCCMCWN2kQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCWAAGx5kwYBVwSCBTDz9piJxpSrwyy2mQcq22Soe+bDJFq/xnc/IWm9ttjyrq7vf/p56VqzdrYnngNwNZ6rbCbCcwBPnT4kY41LV9mo72+f76PBRnbeIVLrLNlVXn5LCkMbZwfCC0F87+ZPdxfhSUAEsJQXRgQBLO35tXaPAFSmyR0AfwugskjmRRCACnIEgABUFsm8CAJQQY4AEIDKIpkXQQAqyBEAAlBZJPMiCEAFOQJAACqLZF4EAaggRwAIQGWRzIsgABXkCAABqCySeREEoIIcASAAlUUyLxJBAOZnJBACEFAnEPy3AOqdUBACEDAngADMkRMIAT8EEICfWdAJBMwJIABz5ARCwA8BBOBnFnQCAXMCCMAcOYEQ8EMgWAB5eh/AruEDUpPO+mSgPeu2y2Cxb3YTqx8ekplPe+qkr+KWG2Vg3+stRzpRHZUXvj3cSceUSqFH9g/dHXQmBNCEbb4HgRBA0G5FvwgBLDwCBIAALhPgDmDhHxav38EdgNJkuAPgvwBKq2RaBgEo4UYACEBplUzLIAAl3AgAASitkmkZBKCEGwEgAKVVMi2DAJRwIwAEoLRKpmUQgBJuBIAAlFbJtAwCUMKNABCA0iqZlokigDx9NuCLZ45IPe2sJwEfWXOLVAq9s4taO3ZUvn/zFdPFzTqssH6DlB97tiXmq9qEvDF6LOto0/rlQkkeX3NbUGbwg0BBaVwEAQi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEkAAtrxJg4ArAsEC+KI6JvUO+xv5DaWV0psU/8+AhkUkdTW4xTezTkS6Z8uk42NSH/5y8WUdVUh6V0hx89aWjibTugxPjTvqcvGtFKVLNi8bCCoULIA8vRAklT0i0ggC7PWiRB4VkdWz7fHnwF4ntXBfUZ4DQAALD8bzdyAAz9O5ut4QwNXxmve753sSkDsAJcDGZXgj0MLA+S9AEyMEwKPAC//I+PsO7gCUZoIAEIDSKpmWQQBKuBEAAlBaJdMyCEAJNwJAAEqrZFoGASjhRgAIQGmVTMsgACXcCAABKK2SaRkEoIQbASAApVUyLRNFAAcnTsrF6brpQbMOu6vv2paXZPyY93dJO+xR4ER+JSI/vhCkMXxSLn3wXtaITesXVl8jPTvub8kcbVyU9y+cMu0j67DlXUXZUd4UFBP8HEBQGhdBAAKuCCAAV+OgGQjYEkAAtrxJg4ArAgjA1ThoBgK2BBCALW/SIOCKAAJwNQ6agYAtgWAB5OnXgO+On5BGh/0a8J7yRunrKjVt21lJ5RPb7cs4LZGyiNzalnJeUvk442Tb8oksE5Hbg0KDBZCnF4LsGj4gtQ57/dmeddtlsNjXtDSfSipvBy2R34sGJZGH29o7Jam85rfloM76JZEng65EAE3Y8vzRYCIIIOgnyMVFCEBlDAiAOwCVRTIvggBUkCMABKCySOZFEIAKcgSAAFQWybwIAlBBjgAQgMoimRdBACrIEQACUFkk8yIIQAU5AkAAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5kQgCyNOHg35WHe24NwJtKq2SUlJoWtVJERkxX91sA2cedR5si6iKyDfZxppXn5nj+qDU4CcBg9K4CAIQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCWAAGx5kwYBVwSCBVA//k9Ja1OuDrPYZoo3/EKS5c1/I//finn4LcDE9JScrk0sFqGr63uSomwsreS3AD8xlWABnHvoXpk+d8bVwBfbzKq9r0r31pvnlMnDC0E+mvxaXjp7dLEIXV1/fWmlPLf2zraeeCFIMxAE0EQDASAAVwa74mYiPAjEHcAVT8flN7a/Eow7AJdjusKmEMAVgvrpb+MOgDsAlUUyL4IAVJAjAASgskjmRRCACnIEgABUFsm8CAJQQY4AEIDKIpkXQQAqyBEAAlBZJPMiCEAFOQJAACqLZF4EAaggRwAIQGWRzItEEMDFd96Q6cnvzY+aZWDvb++TrjVr50Tk4bMBT9cuyD8m/50lXvPaA4Ue2d53XVsunw3YDCT4SUDzaRIIAQioE0AA6kgpCIGlQwABLJ1Z0SkE1AkgAHWkFITA0iGAAJbOrOgUAuoEEIA6UgpCYOkQQABLZ1Z0CgF1AsECeOr0IRlrXFJvKGbB+T4ZaGTnHSId9vqzystvSWFo4yxu3gcQc/MWmx3hQSAEsNihxb0eAcTlr5uOAFR4cgfAo8Aqi2ReBAGoIEcACEBlkcyLIAAV5AgAAagsknkRBKCCHAEgAJVFMi+CAFSQIwAEoLJI5kUQgApyBIAAVBbJvAgCUEGOABCAyiKZF0EAKsgRAAJQWSTzIhEEMPM5chONzvpw0Acq22SouzxnfOO7n5C0XjMfa5aB/U8/3/L2o+PVc/LOd59nGWlee233CnmwclNb7oikctC8l2wDl0siO4Migh8FDkrjIghAwBUBBOBqHDQDAVsCCMCWN2kQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCUQLIAvqmNSl2nbbjNO21BaKb1JcU7KZ9VRSSXNON22/KbSKiklhdnQdHxM6sNf2jaRcVrSu0KKm7e2pEymdRmeGs842bZ8Ubpk87KBoNBgAeTphSC7hg9IrcNkt2fddhks9s0uTfXDQ3J+7zNBS+T1ouKWG2Vg3+st7Z2ojsoL3x722nJQX5VCj+wfujvoWgTQhG2+JwERQNBuRb8IASw8AgSAAC4T4A5g4R8Wr9/BHYDSZLgD4L8ASqtkWgYBKOFGAAhAaZVMyyAAJdwIAAEorZJpGQSghBsBIAClVTItgwCUcCMABKC0SqZlEIASbgSAAJRWybQMAlDCjQAQgNIqmZaJIoCDEyfl4nTd9KBZh93Vd61UCr1zYt4dPyGNDnsU+J7yRunrKs2etTF8Ui598F7WiE3rF1ZfIz077m/JHG1clPcvnDLtI+uw5V1F2VHeFBQT/CBQUBoXQQACrgggAFfjoBkI2BJAALa8SYOAKwIIwNU4aAYCtgQQgC1v0iDgigACcDUOmoGALQEEYMubNAi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEkAAtrxJg4ArAgjA1ThoBgK2BBCALW/SIOCKAAJwNQ6agYAtAQRgy5s0CLgigABcjYNmIGBLAAHY8iYNAq4IIABX46AZCNgSQAC2vEmDgCsCCMDVOGgGArYEEIAtb9Ig4IoAAnA1DpqBgC0BBGDLmzQIuCKAAFyNg2YgYEsAAdjyJg0CrgggAFfjoBkI2BJAALa8SYOAKwIIwNU4aAYCtgT+AxBxhcTqHAGHAAAAAElFTkSuQmCC"
__version__ = "v.1.8.0"
__program_name__ = "Tranche Time Analyzer"
sg.theme("Reddit")
if sg.running_windows():
    font = ("Segoe UI", 12)
else:
    font = ("Arial", 14)
sg.SetOptions(font=font, icon=icon, element_padding=(5, 5))
screen_size = sg.Window.get_screen_size()
image_aspect_ratio = 0.5

checked = b"iVBORw0KGgoAAAANSUhEUgAAAGQAAABkCAYAAABw4pVUAAAR7klEQVR4nO2deZAcV33Hv7/3XnfPzB7alYyR0OEjOMfKFOEKxCaMFEJMqlKAtOoNAQpCFRhDAJsbA1ZPKzJHOIwNJI4h4CP42LGkMuWqmFRS3k1SJApHoFISYDkRpdWBDkt7ztH93vvlj96RVtIcu9rZ1e7sfqq2dqXt7u3pb7/f9X79mgCgN3f0Q9JJf9Sa0mowEwDCMvMDkZXKe87ExXsfD14Q0LZg6DYn03WXjopgG2NZi/mHSMBJdaJcGP5r2hYePU1CdVkbM4HE5T65JQkzQwgGW1IAutnGINDy0LhcEFHiKgiTI2JZi4WCarQBgy0YTMCyOZsFDNjEDNV3C3UFYbasnLQQyoPRZYC5uWe5VCBAqpS0JoaOC7aeKDUFYbas3DayJnpGF0tfYsJ+ttYIoZZVmQHWahIgwXH0GxDyNsdtf2lcHmMiUdVP1BCErXLSZE10oKj1jU+Ga0/N5UkvEX7gf/nQ42YCA8pte4WOi4YAeeFGNQQhK1VK2cKZzz4ZrjuVDQ6mNuHqaP/+/LL3vwR6enzeu/KAk791Q3FLcCjwnPQ/1tq2uiDMUsdFWCF/Dma6si8fh/lr7IWbBQGLZZHOp6fH5zCkatcqfgpMCkPPxuVRS0JKsGVckG7UderWagIRw+8//xdJeQUhXfyHl0kIAhYXC0MMHJr8Xp2GYe/FcCISgK13nvxtl2hFHBcv7VAthYa1DnnptrFHP9mxLxGDqd7Fr8bMrmIQCITgLTtPrFFE3yHGH1mGlDI9o8O0Jg6kBGxc5L47T/2HNvYvdm+nAwhYoIoJq8WMBAmQAwLQPh0/5HWufm1x7Jglntkd0OowMafSz7vBjJ/anQ0OvmIwhzLC6ZdCpp19V2zi/+Dob5JUry2OHTdJ4rlUvgAmYk6qTme/QMDU7QhClCZOasdrv34VvN8DEfv9PO3rPO0RUommCPFKEh6T0QLU2gVJBlsCLAOCSAgpFBEJVBJtZgubVDD4vGvBRETEUuhVM/2bM/bEFtLKpBrZsqaKmQ0BwnHbhFCusCaGjia00dEZEI0QcxlEzOB2Al0lnBSZuHTB/UlkIWcchS710Og8mJlBYDfdKU1chtGlHxlbfkqAfkBWHqDS+Em87NrxfB8ZAHj9B57xOp/ftpGtflQq74XGlGc9p7QsyCQMa4V0hVQemai0h2G/kt+++t/r7fPUPddFIPpJb27ouOOtuM6Y8qzzsmVBADBbq5y0YOYzcVy4Zdf2NZOZMFM2gLxyI7hnHzjMTTpxEGcDVoNEessdh7Y4Xuer42jMEuii2tRMWfKCMFsrnZRg5uNxafSmPTuv/Vk2eFpduf8k5/NkBkPosxuHybfJiFO/+fb/fb5x09+wJm6aP13SgjBbFtIBQCVTHnnjnp3X/uzmv2PnvvdQXGcv2r8R5Pv9UqfbHnHc1JqoOGKIxKxHB7CkZwGZhZBGqrSIy8W37/qra/e+7OYfNRADyAaQ+T4yvPHGr7vprs1RaVQ3SwxgSQsC46a7lC6PfGrPjnX5l938I+fH9728gRisBkPSvcHQB91M9y1R4bQmUFOtzJI0WcxWe22rVKlw+qFdufWfm7zQDcR4Wg2GpLfecegm5bZ9NS6PaYBlsxtEltwIYbbGTa1QcXFkr7Bn3uX7LAdzMPX2CQIWg+Fm7QeHXqi89CPMBtZqceFcRjNYUoIw2EonJXRcOq5NvC0fXh/19IAr0wnVd0rmfl73kWNtUN5uId1uo8t1GxVmwxIShFmQZAJZo4t9e8L1h/1+ltVm96ZA2RxkGJJd0WEfcLyOF8XlsaY68QtZQoLAOKkVMoonPrQ73PCv2YBVpQRSi2zAcjAkvS13OHDbVvVGxTOaSMyp310SgjCzdtMrVVQ8/cCe3IavVaKlevtUnPiWOw5uUV5nLiqe0ajSJdJsWl4QZjaO166i4sjP2mx0i+/3T9uJv+kzB3/H8ToeMKZk2Ro5F078QlpbELYslEPWxGMU6z97ILym1NPjT8uJ+8HxdtfNPC6k6rA6rtnY1mxaWxAhjHLSwsQT78rvXPfLbPC0auDEUXHiEPo7KtXRE5fHNdHsi4bTpWUTQ2arvcwVqjx+6mu7wqv6p+c3km22bh+63cms2hYVnptzJ34hLTlCmK1xvE4VFc/8cNUL1n7E72+c/Pn9SUS19Y5DN7mp9s/GxeF5ceIX0nqCMLNQLlkTjZGJ33Lfeyju2Vc/+QsCFvk+Mn5wZINy09+1VrNlMyeZeCNaTxAio5yM0HHhlny44dnGfiMpp2cDVizFY9LxVs1lJt6I1hKEWbvpbhUVhv9+d7jh4cQnbG7gN5Jy+hV0+G4v3f2qZpfTZ0rLCMJsrXQzKi6N/KIM+uB0ioaV5G9bMPR2N73yfeXC6Xl34hfSIoIwC6GY2caAfduT4doCkEc9v+H7/XIw3Ky3BIeuF076XhNNGDBftpFRoUUEgXHSK6SJip/Kb3/Bj7MBq3y+r/boOFvB/Wmbku5jQqq0MREthMa/RS8IszVOaoWKCmf+eVe47ksVM1Rvn2wOMp/vMx0dV9zrpFZMJn9iQVyLxZ0YsmWhPNJR4Yyx7jsBpk2AHayzy7lp2MO3eJmVb4uKl99vTGVB3BWXDCWlEdal9+0Jrzjs96PKQzLnSPwG6d4dh18i3fTdujxmwPOf/NVj0QrCbI2b7lJRafi7j4cbHm04v8FMPT0++8HxdkHOI0TCNSZeEH5jKotSkMpUbFwaG9K27f1BwKJhiDtZNLTC3Ot4nb9lookF4zemsuBOaDoQCSuEQ9qUb34i7B7evxFUL8StOPre3NC7vXTXW6PiaY0F5DemsggFYe2mu1RcGv3bPeGGpxqZqqn5hnTS9+jy+ILzG1NZkHdJLRhslUqruDjybLpdfcz3+2U+B1Ppub14BybkgXf0HEwVhPOwECoVx+N2viabLoVFNUIEBIMI1sTv/oePrZkA/LrZeDaX1KnGhfNVJ9P1ooWUb9RiQZ/cVJitcTJdMi6N3bMrXD+QZOO1TdXZOlV42HdTXe+JCnPfMdIMFoUgDLbKzYioMHwgInl74hdqR1WVJoWtnz56FUn3Ph0XLdguWL8xlbkXhJkBaIA1AM3MdcPTakcQIAaI2OqbnwzXFuD7qP1APtMABgQzk3D5QaXSXVaXGQvYb0xl7oYwMzPYCulK5WYUkQCYwWDEpZGL1viocyDjZFaq0vipb+wO1w9kg6dVvq92rSobQA6Gm3UvHwpSHVe+pnwZ5sVnw5yMEGZrSDrkZVZKMD+ny+OPlQvDH47Lo++MiiMPSpUiJOuk1H3yiMFWqpSMCyMHJdxPJj1Vm+qEuJPz4uGRG1WqfXtcGtbNeMxsPmnyncNghvHS3VJHhZO6NHaXlJlvP/rpjuNTNrrf3/Hrnzpu21ei0miDEJSYhBIGpffuDteO+/0sa0ZVzNSTA7/hCyc7ZFk/SAxhrAZhcZiqCk0UhAEi66Y6ZRxNPIxi6eP5z244AiQdHSf2JQ9SjB/7MeW3r75rW+7INsfruCGOxk21u5jZGi+zUpYLz92/O7f++w1NVVIa0b25w193MquuLU88Z0hcvqnYS6VpghCEkU5a6vLYx/Pb13wRmCx152DydC48zQasACYWv76bhLwBjIueeZmsVVFcGv21YP5IELAIkatdxe1nme8jvSU49GY33fX2qHhGk1g8fmMqTfEhzGxUqlOaqPDF/PY1X8wGrJLQk/SFJiYpAhJrk/qnqDRyWkhHTkZiUw4IliolrNG35sMNp/dvBCEMqwqStPDAbgmG1imV+hsdFy0w/bVFFhqzP3FmltKRUWn0JKzcEQQsBgFbc14iWYxFPhF2DxPo+8ptY0wZQZNldRkVR763K1zXP2mqajryAQwIgFgQfVu5mW6ry4xFvEL3rIc1E1mhPGnK48/mwzXjYK5beQWAE/sGCADY2oeZzZ8zWzEZFrOQLul4YlTJ1PvBTJtytWcAk2x8s94WDN3mtq16XbQAukZmSxPuJCZmAwJ1+35/7ShoCoPhJgMwiVH7L3Fp9JB0UgLJQsPG8ToER6XPPHbHqqFsbqDmE06VbLx3x7GNwkl/Pi6NGlyG1s9mM2tBCCCrIzDhKrz4VasBIAiCBsclzgaQ+bs2FAF6UDltYGtj5barqHB670Za/42k13ZzDVN1rtuQGPcL6XjWaFyO1s9m04QRQmTZGMdrT8M4vwsA+zfmGl6YTchZgElLvi8ujRalk3KtiZk1vzcMySKfB2okjtlgQOb7yKzC4e1upvvlOhqb10cG5pKmOD8CWAgHFnYzAFRyjnqEYWj9fogn7lg3xDb+gpvpYBMXb9u1c91/+z7LWn1VlQmn3nDolcrNfDounjFYZNl4PZoT9gLCmjKI6Q8BpnqV2Knk+5KVO6ln7c7y+MmeXeH6e4IgEDXL6swE+HhHcDAlyPk2kRCWDbWCqarQpBFCQkdFJiFe9KbgyHUAcRBMKxdggDjfRya/fd0vk1V2qucbAJDNDch8nsw4qzud9IoeExc0YWFPOM2UJn4Ya5xUp1JENwGggZkdm6ovPHyOpHC4WW8Njr5Gpdo+HBeGW8pUVWiaIMwgtgYgegMA3gTMZHW1qstzTzk4AcDNwZGMFOJbAGBhW/LlZU0ThAhCRxNg0I3+zkNrw5AsGoa/06Niqk4Rf87JrLjOxEXdqu/LauKHImI22kuvSLN1/wRgyiI36+NXTJUfHs06btsH40JrRVUX0tS77KzZstoHaKZmq9oBCcjjT4MjGQZ9E2BYtPZ7FpsqyFmzRfIPtgSn1oUh2cZZe20SU9VnPDZ3upmu63QLm6oKTf5w58yWoNIbAWDgEs1WxVT1BsdvUF7brXFhuOpEVqvR9LuNAbImBoHeAgCDl2S2kqjqA3ezB2G/SSTIwrS0qarQdEEIJHU0zqTcV23ZcfzFCMG+3z+jOzsbJFHV4dNDgZfp7tEtmADWYq4+pHHcNiFtdAtAnPRRTY+ztaodh1/iuG0fi4vDLVFWny5zJYiMS2NMwnmLHxxf3bMvN81SCtOJnueR77MkpvuEUMpag1aqVTVijgQhYtbGSa3otFz+RBiGdj/21ZnJYwoCFq//wAF3MNyszcahD7uZlS/X0fyuxLMQmMvpThmVhq3y2t63JfjVQ/nw6p9UnHUQTHXOOYQh2TAEAyj3BodvECq1o1VmAGfKHApCxGwZgKtk5nE/OPZmYN9P8+H10eTFP4vfzxLPnrgGMfWRpNvBNmXM/C0atpCY04YAAgkdF1k66Wssm71A9//15o6eIaAAQCN5e41nf37sShBd7aU73bg0Bmv1khQDuARBWNCM3q5DJMjokiWQkCp1bbKuiziXUTCDrYExZZQLw4aS1wu1gBjMTDPPm6YtSE+PzwBAVo6xtcRA1Xe5ViMpdzBrXeLkMQI+J2lyykQAtYwDJzAzkyQxOtNdpy1IUpdiMQD84nnm6A9T7Ve8ojR+Us88JOVq/+TzVVrUsJfpUlFx+Fdk1X+BmfIEe34gU5sZm6zBkPTWHSfeqsvjjzpex0tBlXHZAlZmVlTuLIaOys8Y0Nv2hM8fD8AiROWtn42ZkSCV14nu3k4Hstnglav/+Nbft8Z0AwZJs8HSRRAz4IGkGuPnfvGfe+66oQhmmun7gusKwlZXucjEQRCIMAw1BsN/m9lpLx2CgEU1MYyJiFTtV9XWFUQqNTnU8uf9f9IZwuT7EJh+mWrJkO+r1WzO5HgnrDZxTWtSVRAGrHTS0salawDaO/ZqVkEPV/O4jH2XfN4tS+LAz/cZ+wEFUKT10Hon1SHiaKLqQpvVRwiBrImYpPqE/6EfPJG/lYpPzdHJLyEiABDS/cxkAFQ1pKRt4dGqv2Cwddx2YXR5L1lzJ5N5BsYaobyWiE3nC6vLJCCEVs7VAvRRqbzXxdF4zWVoa/oQAok4GreO0/ZKKPpeHI1bkOSky3yZaUMSlghKupJIoJ4YQAOnTiARxxM2+VmIxbHuw8JExwULgBv1BTTMQ86pabk1EunLw3S7ZSYFqfIo7MWHXNKJ33whGDxOQtmLnoRdZj5hECyRgADElx2vQyy0xSCXFgw33S2std8iMJO/49gXhJP6S6PLqclO82Vx5gsiK6UXa11+aKJr9fv/H4FVCYlw1NLbAAAAAElFTkSuQmCC"
unchecked = b"iVBORw0KGgoAAAANSUhEUgAAAGQAAABkCAYAAABw4pVUAAAFqUlEQVR4nO3dT2hcVRTH8d+5976ZiU3/2z+CKLqQElyo2Eq1UKjoXslERKHuu6jVggvR1wctgkW6dCUoutCOpSpddKkI0qo7a1qCUCw2lja00kkzf96797hIBpv0ZTIvNekhcz4wECYzw818575332RxCQCq8fhbplw56LPWFjATAIJaDgwiNrZ8PWStj7+Ot8Y0nFx+M6qsPZa1p8AhhbZYfkQGUWUNWo0bR2k4uXyDTLQ2hJQJZO714PoSM8MYBrNxAK3jkIJAOjXuFSKaOVXAAGA9TMnhsEANBgcwmKbjqUViIEwfhrqfFlzXF+HALhowxpURshYY/P+Osm8QrCvb4FNk6VToFmXeIMyBo9Ig+aw15n16NLA/76z1aZouzZhXqAgRvCPDrdajgD0QlQafSlt1JjK5R6bcIAwOLhog79Oxhr/x3Klk28TSDrsv/FQ9cOmEX4/vo9Kq7Wna8ATYuQ/KDUKgYKOKS29NfHAq2TaxO75Y2Tz6i06Nu1Df9YSr7X+o8XJy+X1rBk7P97j8QxazzdoNkCmdBzNtHqmltdqIn/uwOGYzOlrTJdpthoaqnCTEwOwTbjzEfBpMkbn6R9q8GchYCw6MOZcbXU/qAAAiRvX47Ptm1swJUbjbP2ClimM2STL3/SEO2SUC2XlXRwsHuQMTiBgAXj1ybZtnv8575hCyvp4pxji2lggUTR5/d+O56RhMABVamhYLEscGCfilw1cfcKBPPeN5wFlrAeuiQi+14nTe9uB55PDEmXbafOObhMbyZ8r8CgWJcQiIQeey9Ivymq17GvW/A3GxT8BKx8RcWbVpp/etE3vji9uTQ2gh4Z6PHj1ffXdK/4bxx4x1exr1K376wlNvt98IxjQnr2VRafDxhivvABFXj/f+Pvf8wM5qisAbyFgmhpm7QlAdRETEHHhj0WcW/n4qgAP028geEAWg8CpUvzAURoMIo0GE0SDCaBBhNIgwGkQYDSKMBhFGgwijQYTRIMJoEGE0iDAaRBgNIowGEUaDCKNBhNEgwmgQYTSIMBpEGA0ijAYRRoMIo0GE0SDCaBBhNIgwGkQYDSKMBhFGgwijQYTRIMJoEGE0iDAaRBgNIowGEUaDCKNBhNEgwmgQYTSIMBpEGA0ijAYRRoMIo0GE0SDCaBBhNIgwGkQYDSKMBhFGgwijQYTRIMJoEGE0iDAaRBgNIowGEaZwEDZEgO5SvDBmpuI7EfUcZGioygBAwdYRAvEitvPpG8TMABFzvehTe942L0koIGZzHbiwyY//Whm8/+nm5LVMNwabg5jLA+tde+qfPw1HZ8FMNUKI495mS8F9DIEkoeyV+K/X0ubkl1F59ZNEBN0jrIPBDGRpa8wTvX7y0JbJGGwSdHb9XFihIJ3tRL9KaGz37njH1hff3snc2pD5NnJ2ou4zHs6WwORu8sSFMyePPdsAMxXdL7hrkPz9bYnjODZJkmT4Ifmx2KD7RxyzyYvhfZvIDcz7vK5BXORmplpt1v1JkgSAqVqFQXWxQ165aiMI+fvfMkXlqyHz6byHr/wgRN5GAy5rNB8B6Of6LnbxEOctdRm/L3bYK9f0CXz2OWMUcAC1Qxh/MCoPmrR9KxDojlVubhAGm+DbDBu9sze++O1n+6l5eqlG3z/aAMBE780sgnKv5Wg4Gc/9BSOEqLTa+Kx51odwxJIbgw8eaC3dkFekMmBTGzw9bKw7aKPKC2l7Mnd2AF3OIQRj0vZkiKJVz1hH36XtyQAD1tVUURnARK58nyEy6BYDWOCkTiCTprfC9M/G6OXG4mXpVADABOr6iXaYPpbN+1b/VzOwfoO1eN1mxe0cAFqgSecldX4sA8PMt8i4AM5d1qrlQghEBgbAR1F5tQHpDLh3mEsD603g7BMCMw0nVz60UXmf960ymPWfVsuJKFhbykKWfl5ft2Xfv1QjGxJX1GgkAAAAAElFTkSuQmCC"

# results queue for threads
results_queue = queue.Queue()
cancel_flag = threading.Event()

weekday_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
analysis_options = {
    "weekday_exclusions": [],
    "put_or_call": True,
    "idv_weekday": True,
    "news_exclusions": [],
}

news_events = {
    "CPI": [],
    "Initial Jobless Claims": [],
    "Retail Sales": [],
    "ADP": [],
    "JOLT": [],
    "Unemployment/NFP": [],
    "PPI": [],
    "GDP": [],
    "PCE": [],
    "Triple Witching": [],
    "Beige Book": [],
    "ISM Manufacturing PMI": [],
    "ISM Services PMI": [],
    "S&P Global PMI": [],
    "Fed Chair Speech": [],
    "FOMC Minutes": [],
    "FOMC": [],
    "MI Consumer Sent.": [],
    "Chicago PMI": [],
}


# Allows custom checkbox icon
class Checkbox(sg.Checkbox):
    elements = []

    def __init__(self, *args, **kwargs):
        font = kwargs.get("font", sg.DEFAULT_FONT)
        w, h = sg.Text.char_width_in_pixels(font), sg.Text.char_height_in_pixels(font)
        text_color = kwargs.get("text_color", sg.theme_text_color())
        background_color = kwargs.get("background_color", sg.theme_background_color())
        self.images = [self.icon(i, h, text_color, background_color) for i in range(2)]
        size = kwargs.get("size", None)
        size = (size, 1) if isinstance(size, int) else size
        size = (
            (size[0] * w + h + h // 3, size[1] * h)
            if size
            else (len(args[0]) * w + h + h // 3, h)
        )
        kwargs["size"] = size
        super().__init__(*args, **kwargs)
        Checkbox.elements.append(self)

    @staticmethod
    def initial(window):
        # Called after window finalized
        for element in Checkbox.elements:
            element.widget.configure(
                indicatoron=False,
                image=element.images[0],
                selectimage=element.images[1],
                compound=sg.tk.LEFT,
                borderwidth=0,
                offrelief=sg.tk.FLAT,
                selectcolor=sg.theme_background_color(),
            )
        window.refresh()

    def icon(self, i, h, fg, bg):
        box = (h, h)
        im = Image.new("RGBA", (h + h // 3, h), bg)
        im_check = Image.open(BytesIO(base64.b64decode(checked if i else unchecked)))
        im_check = im_check.resize(box, resample=Image.LANCZOS)
        im.paste(im_check, (0, 0), im_check)
        photo_image = ImageTk.PhotoImage(im)
        return photo_image

    @staticmethod
    def clear_elements():
        Checkbox.elements.clear()


def with_gc(func):
    """
    Decorator to garbage collect threaded functions.
    This resolves the Tcl_AsyncDelete error
    """

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        result = func(*args, **kwargs)
        gc.collect()
        return result

    return wrapper


def analyze(
    df: pd.DataFrame,
    settings: dict,
) -> Tuple[
    pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame
]:
    short_avg_period = settings["-AVG_PERIOD_1-"]
    long_avg_period = settings["-AVG_PERIOD_2-"]
    short_weight = settings["-PERIOD_1_WEIGHT-"] / 100
    long_weight = settings["-PERIOD_2_WEIGHT-"] / 100
    calc_type = settings["-CALC_TYPE-"]
    agg_type = "M" if settings["-AGG_TYPE-"] == "Monthly" else "W"

    def calculate_avg_pnl(df: pd.DataFrame) -> float:
        if df.columns[0] == "Date Opened":  # OO BT data
            return df["P/L"].sum() / df["No. of Contracts"].sum()
        elif df.columns[0] == "TradeID":  # BYOB BT data
            df["P/L"] = (
                df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
            ) * 100
            return df["P/L"].mean()
        else:
            raise ValueError("Unknown dataset type")

    def calculate_pcr(df: pd.DataFrame) -> float:
        if df.columns[0] == "Date Opened":  # OO BT data
            return df["P/L"].sum() / (df["Premium"] * df["No. of Contracts"]).sum()
        elif df.columns[0] == "TradeID":  # BYOB BT data
            df["P/L"] = df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
            return df["P/L"].sum() / df["Premium"].sum()
        else:
            raise ValueError("Unknown dataset type")

    def calculate_rolling_averages(df, short_avg_period, long_avg_period, agg_type):
        if agg_type == "W":
            short_avg_period = int(short_avg_period * 4.33)
            long_avg_period = int(long_avg_period * 4.33)
        short_avg = df.rolling(short_avg_period, min_periods=1).mean()
        long_avg = df.rolling(long_avg_period, min_periods=1).mean()
        weighted_avg = short_weight * short_avg + long_weight * long_avg
        return weighted_avg

    def create_output_labels(df, long_avg_period, start_date, end_date, agg_type):
        output_labels = pd.DataFrame(index=df.index)
        for i, (date, row) in enumerate(df.iterrows()):
            if agg_type == "M":
                current_period_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
                previous_period_start = (
                    current_period_end - pd.DateOffset(months=long_avg_period - 1)
                ).replace(day=1)
            elif agg_type == "W":
                current_period_end = date.to_timestamp() + pd.offsets.Week(weekday=6)
                previous_period_start = current_period_end - pd.DateOffset(
                    weeks=int(long_avg_period * 4.33)
                )
            else:
                current_period_end = date.to_timestamp() + pd.offsets.DateOffset(
                    freq=agg_type
                )

            if i == 0:
                date_range_label = f"{end_date} - {previous_period_start.date()}"
            elif i == len(df) - 1:
                date_range_label = f"{current_period_end.date()} - {start_date}"
            else:
                date_range_label = (
                    f"{current_period_end.date()} - {previous_period_start.date()}"
                )
            output_labels.loc[date, "Date Range"] = date_range_label
        return output_labels

    def perform_analysis(df_grouped):
        if calc_type == "PCR":
            df_calc = df_grouped.apply(calculate_pcr, include_groups=False)
        elif calc_type == "PnL":
            df_calc = df_grouped.apply(calculate_avg_pnl, include_groups=False)
        else:
            raise ValueError("Invalid calc_type. Expected 'PCR' or 'PnL'.")

        if isinstance(df_calc.index, pd.MultiIndex):
            df_calc = df_calc.unstack(level=-1)

        weighted_avg = calculate_rolling_averages(
            df_calc, short_avg_period, long_avg_period, agg_type
        )
        one_month_avg = df_calc.rolling(
            1 if agg_type == "M" else 4, min_periods=1
        ).mean()

        weighted_avg.sort_index(ascending=False, inplace=True)
        one_month_avg.sort_index(ascending=False, inplace=True)

        if isinstance(weighted_avg, pd.Series):
            weighted_avg = weighted_avg.to_frame()

        if calc_type == "PCR":
            weighted_avg = weighted_avg.apply(lambda x: round(x, 4))
            one_month_avg = one_month_avg.apply(lambda x: round(x, 4))
        elif calc_type == "PnL":
            weighted_avg = weighted_avg.apply(lambda x: round(x, 2))
            one_month_avg = one_month_avg.apply(lambda x: round(x, 2))

        output_labels = create_output_labels(
            weighted_avg, long_avg_period, start_date, end_date, agg_type
        )
        one_month_avg_labels = create_output_labels(
            one_month_avg, 1, start_date, end_date, agg_type
        )

        df_output = pd.concat([output_labels, weighted_avg], axis=1)
        df_output_1mo_avg = pd.concat([one_month_avg_labels, one_month_avg], axis=1)

        return df_output, df_output_1mo_avg

    # get list of news event dates to skip.
    news_date_exclusions = []
    for release, date_list in news_events.items():
        if release in settings["-NEWS_EXCLUSIONS-"]:
            news_date_exclusions += date_list

    # filter df for news exclusions
    df = df[~df["EntryTime"].dt.date.isin(news_date_exclusions)]

    # filter for weekday exlusions
    df = df[~df["Day of Week"].isin(settings["-WEEKDAY_EXCLUSIONS-"])]

    if is_BYOB_data(df):
        df_grouped_combined = df.groupby(
            [df["EntryTime"].dt.to_period(agg_type), "Time"]
        )
        df_grouped_puts = df[df["OptionType"] == "P"].groupby(
            [df["EntryTime"].dt.to_period(agg_type), "Time"]
        )
        df_grouped_calls = df[df["OptionType"] == "C"].groupby(
            [df["EntryTime"].dt.to_period(agg_type), "Time"]
        )
        start_date = df["EntryTime"].min().date()
        end_date = df["EntryTime"].max().date()
    else:
        df_grouped_combined = df.groupby(
            [df["Date Opened"].dt.to_period(agg_type), "Time Opened"]
        )
        df_grouped_puts = df[df["OptionType"] == "P"].groupby(
            [df["Date Opened"].dt.to_period(agg_type), "Time Opened"]
        )
        df_grouped_calls = df[df["OptionType"] == "C"].groupby(
            [df["Date Opened"].dt.to_period(agg_type), "Time Opened"]
        )
        start_date = df["Date Opened"].min().date()
        end_date = df["Date Opened"].max().date()

    df_output_combined, df_output_1mo_avg_combined = perform_analysis(
        df_grouped_combined
    )
    if df[df["OptionType"] == "P"].empty or not settings["-PUT_OR_CALL-"]:
        df_output_puts, df_output_1mo_avg_puts = pd.DataFrame(
            columns=["Date Range"]
        ), pd.DataFrame(columns=["Date Range"])
    else:
        df_output_puts, df_output_1mo_avg_puts = perform_analysis(df_grouped_puts)
    if df[df["OptionType"] == "C"].empty or not settings["-PUT_OR_CALL-"]:
        df_output_calls, df_output_1mo_avg_calls = pd.DataFrame(
            columns=["Date Range"]
        ), pd.DataFrame(columns=["Date Range"])
    else:
        df_output_calls, df_output_1mo_avg_calls = perform_analysis(df_grouped_calls)

    return (
        df_output_combined,
        df_output_1mo_avg_combined,
        df_output_puts,
        df_output_1mo_avg_puts,
        df_output_calls,
        df_output_1mo_avg_calls,
    )


def create_excel_file(
    file,
    settings,
    open_files,
) -> dict:
    calc_type = settings["-CALC_TYPE-"]
    short_avg_period = settings["-AVG_PERIOD_1-"]
    short_weight = settings["-PERIOD_1_WEIGHT-"] / 100
    long_avg_period = settings["-AVG_PERIOD_2-"]
    long_weight = settings["-PERIOD_2_WEIGHT-"] / 100
    top_x = settings["-TOP_X-"]
    weekday_exclusions = settings["-WEEKDAY_EXCLUSIONS-"]
    result = load_data(file, weekday_exclusions)
    if result:
        df, start_date, end_date = result
    else:
        return

    # path and orginal filename
    path = os.path.join(os.path.dirname(file), "data", "heatmaps")
    org_filename = os.path.splitext(os.path.basename(file))[0]
    os.makedirs(path, exist_ok=True)

    # Create filename
    filename = os.path.join(
        path,
        (
            f"{org_filename}-TWAvg({calc_type})_{short_avg_period}mo({short_weight * 100:.0f})-{long_avg_period}mo({long_weight * 100:.0f})_{start_date} -"
            f" {end_date}.xlsx"
        ),
    )

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:

        # Get the xlsxwriter workbook
        workbook = writer.book

        # get the sheets for day of week
        day_to_num = {
            "Monday": 1,
            "Tuesday": 2,
            "Wednesday": 3,
            "Thursday": 4,
            "Friday": 5,
            "Saturday": 6,
            "Sunday": 7,
        }

        days_sorted = ["All"]
        if settings["-IDV_WEEKDAY-"]:
            # This gets the unique days of the week from the DataFrame, then sorts them based on the numerical value
            days_sorted = days_sorted + sorted(
                [
                    d
                    for d in df["Day of Week"].unique()
                    if d not in settings["-WEEKDAY_EXCLUSIONS-"]
                ],
                key=lambda day: day_to_num[day],
            )
        df_dicts = {"Put/Call Comb": {}, "Puts": {}, "Calls": {}}
        for day in days_sorted:
            # check for cancel flag to stop thread
            if cancel_flag.is_set():
                return
            if day == "All":
                _df = df
            else:
                _df = df[df["Day of Week"] == day]
            (
                df_output,
                df_output_1mo_avg,
                df_output_puts,
                df_output_1mo_avg_puts,
                df_output_calls,
                df_output_1mo_avg_calls,
            ) = analyze(_df, settings)
            # store the results and the original df in case we need it later
            df_dicts["Put/Call Comb"][day[:3]] = {"org_df": _df, "result_df": df_output}
            df_dicts["Puts"][day[:3]] = {
                "org_df": _df[_df["OptionType"] == "P"],
                "result_df": df_output_puts,
            }
            df_dicts["Calls"][day[:3]] = {
                "org_df": _df[_df["OptionType"] == "C"],
                "result_df": df_output_calls,
            }

            # create the sheets
            df_output.to_excel(writer, sheet_name=f"P-C_Comb_{day[:3]}", index=False)
            df_output_1mo_avg.to_excel(
                writer, sheet_name=f"P-C_Comb_1mo-{day[:3]}", index=False
            )
            if settings["-PUT_OR_CALL-"]:
                df_output_puts.to_excel(
                    writer, sheet_name=f"Puts_{day[:3]}", index=False
                )
                df_output_calls.to_excel(
                    writer, sheet_name=f"Calls_{day[:3]}", index=False
                )
                df_output_1mo_avg_puts.to_excel(
                    writer, sheet_name=f"Puts_1mo-{day[:3]}", index=False
                )
                df_output_1mo_avg_calls.to_excel(
                    writer, sheet_name=f"Calls_1mo-{day[:3]}", index=False
                )

        # use All df from Put/Call Combined for row and col lengths
        df_output = df_dicts["Put/Call Comb"]["All"]["result_df"]
        # Set the PCR columns to percentage format
        percent_format = workbook.add_format({"num_format": "0.00%", "align": "center"})
        top_x_format = workbook.add_format(
            {"bold": 1, "font_color": "#FFFFFF"}
        )  # white
        for row in range(
            2, len(df_output) + 2
        ):  # +2 because Excel's index starts from 1 and there is a header row
            for worksheet in writer.sheets.values():
                # Apply a conditional format to the PCR cells in the current row
                worksheet.conditional_format(
                    f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                    {
                        "type": "3_color_scale",
                        "min_color": "red",
                        "mid_color": "yellow",
                        "max_color": "green",
                    },
                )
                # Format top x values in bold white text
                if top_x > 0:
                    worksheet.conditional_format(
                        f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                        {
                            "type": "top",
                            "value": top_x,
                            "format": top_x_format,
                        },
                    )
                if calc_type == "PCR":
                    worksheet.set_row(row - 1, None, percent_format)

        # Adjust the column widths
        for column in df_output:
            column_length = max(
                df_output[column].astype(str).map(len).max() + 1, len(column) + 1
            )
            col_idx = df_output.columns.get_loc(column)
            for worksheet in writer.sheets.values():
                worksheet.set_column(col_idx, col_idx, column_length)

    # open file in excel
    if open_files:
        try:
            if platform.system() == "Windows":
                os.startfile(filename)
            elif platform.system() == "Darwin":  # This is the value returned for macOS
                subprocess.Popen(["open", filename])
            else:
                subprocess.call(("xdg-open", filename))  # linux
        except:
            pass

    return df_dicts


def chunk_list(input_list, chunk_size=4):
    return [
        input_list[i : i + chunk_size] for i in range(0, len(input_list), chunk_size)
    ]


def format_float(value):
    """This will remove decmal from float for GUI
    display purposes, but only when decimal is .0"""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


@with_gc
def get_pnl_plot(results, filename):
    table_data = []
    plt.figure(figsize=(8, 4))
    for strategy, df in results.items():
        plt.plot(df["Date"], df["Current Value"], label=strategy)
        # Calculate summary statistics for the strategy
        final_value = df["Current Value"].iloc[-1]
        max_dd = df["Max DD"].max()
        dd_days = df["DD Days"].max()
        initial_value = df["Initial Value"].min()
        total_return = (final_value - initial_value) / initial_value
        win_streak = df["Win Streak"].max()
        loss_streak = df["Loss Streak"].max()
        # CAGR
        start_dt = df["Date"].iloc[0]
        end_dt = df["Date"].iloc[-1]
        years = (end_dt - start_dt).days / 365.25
        cagr = ((final_value / initial_value) ** (1 / years)) - 1

        if max_dd:
            mar = cagr / max_dd
        else:
            mar = float("inf")

        # Group PnL by month
        df["YearMonth"] = df["Date"].dt.to_period("M")
        monthly_pnl = df.groupby("YearMonth")["Day PnL"].sum()

        # Calculate largest and lowest monthly PnL with their corresponding dates
        largest_monthly_pnl = monthly_pnl.max()
        largest_monthly_pnl_date = monthly_pnl.idxmax().to_timestamp()
        lowest_monthly_pnl = monthly_pnl.min()
        lowest_monthly_pnl_date = monthly_pnl.idxmin().to_timestamp()

        # Format the date strings
        largest_monthly_pnl_str = (
            f"{largest_monthly_pnl:,.2f} {largest_monthly_pnl_date.strftime('%b%y')}"
        )
        lowest_monthly_pnl_str = (
            f"{lowest_monthly_pnl:,.2f} {lowest_monthly_pnl_date.strftime('%b%y')}"
        )

        # Create row for Table
        row_data = [
            f"{strategy}",
            f"{final_value:,.2f}",
            f"{(final_value - initial_value):,.2f}",
            f"{total_return:,.2%}",
            f"{cagr:.2%}",
            f"{max_dd:.2%}",
            f"{dd_days}",
            f"{mar:.2f}",
            f"{win_streak}",
            f"{loss_streak}",
            largest_monthly_pnl_str,
            lowest_monthly_pnl_str,
        ]

        table_data.append(row_data)

    plt.title("P/L Walk Forward Test")
    plt.xlabel("Date")
    plt.ylabel("Current Value")
    plt.legend()
    plt.grid(True)
    # plt.xticks(rotation=45)
    plt.tight_layout()

    plt.savefig(filename, dpi=150)
    plt.close()
    return table_data


@with_gc
def get_news_event_pnl_chart(results, filename, sum=True):
    # Get list of news events
    events = list(news_events.keys())

    # Initialize a dictionary to hold summed PnL for each strategy and news event
    summed_pnls = {
        strategy: {event: 0 for event in events} for strategy in results.keys()
    }

    # Sum the PnL values for each strategy and news event
    for strategy, df in results.items():
        for event, dates in news_events.items():
            event_dates = pd.to_datetime(dates)
            if sum:
                event_pnl = df[df["Date"].dt.date.isin(event_dates.date)][
                    "Day PnL"
                ].sum()
            else:
                event_pnl = df[df["Date"].dt.date.isin(event_dates.date)][
                    "Day PnL"
                ].mean()
            summed_pnls[strategy][event] = event_pnl

    # Prepare data for the bar chart
    x = np.arange(len(events))  # the label locations
    width = 0.8 / len(results)
    fig, ax = plt.subplots(figsize=(10, 5))

    # Plot bars for each strategy
    for i, (strategy, pnl_dict) in enumerate(summed_pnls.items()):
        pnls = [pnl_dict[event] for event in events]
        ax.bar(x + (i - (len(results) - 1) / 2) * width, pnls, width, label=strategy)

    # Add labels, title, and custom x-axis tick labels
    if sum:
        ax.set_ylabel("Total PnL")
        ax.set_title("PnL by News Event")
    else:
        ax.set_ylabel("Average PnL")
        ax.set_title("Avgerage PnL Per News Event")
    ax.set_xticks(x)
    ax.set_xticklabels(events, rotation=45, ha="right")
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.25), ncol=min(len(results), 4))

    fig.subplots_adjust(bottom=0.3)
    plt.tight_layout()
    plt.savefig(filename, dpi=150, bbox_inches="tight")
    plt.close()


@with_gc
def get_weekday_pnl_chart(results, filename):
    # Filter weekdays based on exclusions
    weekdays = [day[:3] for day in weekday_list]

    # Initialize a dictionary to hold summed PnL for each strategy and weekday
    summed_pnls = {
        strategy: {day: 0 for day in weekdays} for strategy in results.keys()
    }

    # Sum the PnL values for each strategy and weekday
    for strategy, df in results.items():
        for day in weekdays:
            summed_pnls[strategy][day] = df[df["Weekday"] == day]["Day PnL"].sum()

    # Prepare data for the bar chart
    x = np.arange(len(weekdays))  # the label locations
    width = 0.8 / len(results)

    fig, ax = plt.subplots(figsize=(10, 5))

    # Plot bars for each strategy
    for i, (strategy, pnl_dict) in enumerate(summed_pnls.items()):
        pnls = [pnl_dict[day] for day in weekdays]
        ax.bar(x + (i - (len(results) - 1) / 2) * width, pnls, width, label=strategy)

    # Add labels, title, and custom x-axis tick labels
    ax.set_ylabel("Total PnL")
    ax.set_title("PnL by Weekday")
    ax.set_xticks(x)
    ax.set_xticklabels(weekdays)
    ax.legend(
        loc="upper center", bbox_to_anchor=(0.5, -0.15), ncol=min(len(results), 4)
    )
    fig.subplots_adjust(bottom=0.2)
    plt.tight_layout()
    plt.savefig(filename, dpi=150, bbox_inches="tight")
    plt.close()


def get_top_times(
    df_dict, strategy_settings, date: dt.datetime.date = None, top_n_override=0
) -> pd.DataFrame:
    portfolio_mode = "-SINGLE_MODE-" not in strategy_settings
    all_top_values = []
    top_n = 0
    # Iterate over each DataFrame in the dictionary
    for source, _df_dict in df_dict.items():
        strategy = "-SINGLE_MODE-" if not portfolio_mode else f"{source}.csv"
        try:
            settings = strategy_settings[strategy]
        except KeyError:
            # this is probably the separate put/call analysis we need to parse the source file
            strategy = f"{source.split('||')[1]}.csv"
            settings = strategy_settings[strategy]

        agg_type = "M" if settings["-AGG_TYPE-"] == "Monthly" else "W"
        top_n = top_n_override if top_n_override else int(settings["-TOP_X-"])
        calc_type = settings["-CALC_TYPE-"]
        df_orig = _df_dict["result_df"]

        if not date:
            df = df_orig.copy()
        else:
            df = df_orig.loc[df_orig.index == pd.Period(date, freq=agg_type), :]

        if df.index.name != "Date Range":
            df.set_index("Date Range", inplace=True)

        # Extract the first row of values
        try:
            if df.empty:
                continue
            first_row_values = df.iloc[0]
        except IndexError:
            # missing data for this month
            continue

        # Sort the values in descending order and select the top n
        top_values = first_row_values.sort_values(ascending=False).head(top_n)

        # Format values based on calc_type and add to the list of all top values
        for time, value in top_values.items():
            formatted_value = (
                f"{value:.2f}" if calc_type == "PnL" else f"{value * 100:.2f}%"
            )
            all_top_values.append(
                {
                    "Top Times": time,
                    "Values": formatted_value,
                    "Source": source,
                    "RawValue": value,  # Keep raw value for sorting
                }
            )

    # Convert the list to a DataFrame
    if all_top_values:
        result_df = pd.DataFrame(all_top_values)
    else:
        # empty
        result_df = pd.DataFrame(columns=["Top Times", "Values", "Source", "RawValue"])

    if not portfolio_mode:
        # Sort all values based on RawValue and select the overall top n
        result_df = result_df.sort_values("RawValue", ascending=False).head(top_n)

    # Drop the RawValue column as it's no longer needed
    result_df = result_df.drop(columns=["RawValue"])

    return result_df


def import_news_events(filename) -> bool:
    global news_events
    """
    Import CSV downloaded from https://www.fxstreet.com/economic-calendar
    populates the dates for the releases in 'news_events' dict
    """

    def get_triple_witching_dates(
        start_year: int = 2000, end_year: int = dt.datetime.now().year
    ):
        """
        These are not in the calendar and must be calculated
        Triple witching ocurrs on the third friday of March, June, Sept, Dec
        """
        triple_witching_dates = []

        for year in range(start_year, end_year + 1):
            for month in [3, 6, 9, 12]:  # March, June, September, December
                # Get the first day of the month
                first_day = dt.datetime(year, month, 1)

                # Find the first Friday
                friday = first_day + dt.timedelta(
                    days=(4 - first_day.weekday() + 7) % 7
                )

                # Get the third Friday
                third_friday = friday + dt.timedelta(weeks=2)

                triple_witching_dates.append(third_friday.date())

        return triple_witching_dates

    def get_event(name):
        """
        Helper function to add news_event column to the df
        """
        keyword_dict = {
            "Consumer Price Index": "CPI",
            "Nonfarm Payrolls": "NFP",
            "ADP Employment": "ADP",
            "Initial Jobless Claims": "Initial Jobless Claims",
            "Retail Sales": "Retail Sales",
            "JOLT": "JOLT",
            "Unemployment": "Unemployment/NFP",
            "Producer Price Index": "PPI",
            "Gross Domestic Product": "GDP",
            "Personal Consumption Expenditures": "PCE",
            "Beige Book": "Beige Book",
            "ISM Manufacturing PMI": "ISM Manufacturing PMI",
            "ISM Services PMI": "ISM Services PMI",
            "Fed's Chair": "Fed Chair Speech",
            "FOMC Minutes": "FOMC Minutes",
            "Fed Interest Rate Decision": "FOMC",
            "Michigan Consumer Sentiment Index": "MI Consumer Sent.",
            "Chicago Purchasing": "Chicago PMI",
        }

        if "S&P" in name and "PMI" in name:
            return "S&P Global PMI"
        else:
            for keyword, event in keyword_dict.items():
                if keyword in name:
                    return event
            return ""

    # load csv, config dates and filter for US events
    try:
        df = pd.read_csv(filename)
    except Exception as e:
        return False
    if (
        "Start" not in df.columns
        or "Currency" not in df.columns
        or "Name" not in df.columns
    ):
        return False
    df.drop_duplicates(inplace=True)
    df["Start"] = pd.to_datetime(df["Start"])
    df = df[df["Currency"] == "USD"]
    df["news_event"] = df["Name"].apply(get_event)

    for news_event in news_events:
        if news_event == "Triple Witching":
            news_events[news_event] = get_triple_witching_dates()
        else:
            filtered_df = df[df["news_event"] == news_event]
            news_events[news_event] = sorted(filtered_df["Start"].dt.date.to_list())

    return True


def get_next_filename(path: str, base: str, ext: str) -> str:
    """
    Takes a path, base name, and extension.
    Checks if a filename already exists with that filename
    Adds (x) to the filename and returns the complete filename path
    """
    # Create filename
    filename = os.path.join(path, f"{base}{ext}")
    counter = 1
    while os.path.exists(filename):
        filename = os.path.join(path, f"{base}({counter}){ext}")
        counter += 1
    return filename


def is_BYOB_data(df: pd.DataFrame) -> bool:
    return df.columns[0] == "TradeID"


def load_data(
    file: str,
    weekday_exclusions: list = [],
) -> Tuple[pd.DataFrame, dt.datetime.date, dt.datetime.date]:
    """
    Takes a Trade Log CSV from either Option Omega or BYOB
    and returns a dataframe containing the trade data and
    the start and end dates of the dataset
    """
    # Load the CSV file
    try:
        delims = [",", ";", "\t", "|"]
        for delim in delims:
            df = pd.read_csv(file, delimiter=delim)
            if len(df.columns) > 1:
                break
        if len(df.columns) <= 1:
            # none of the delims worked
            raise UnicodeDecodeError
    except UnicodeDecodeError:
        sg.popup_no_border(
            "This does not appear to be a backtest results\nCSV from either OptionOmega"
            " or BYOB.\n\nPlease choose a different file"
        )
        return

    # remove duplicate rows in case human error in combining csv files
    df.drop_duplicates(inplace=True)

    # Determine which type of data, OptionOmega or BYOB
    is_byob = is_BYOB_data(df)
    if is_byob is None:
        sg.popup_no_border(
            "This does not appear to be a backtest results\nCSV from either OptionOmega"
            " or BYOB.\n\nPlease choose a different file"
        )
        return

    elif not is_byob:  # OO BT data
        # Convert 'Date Opened' to datetime format
        df["Date Opened"] = pd.to_datetime(df["Date Opened"])

        # Add Day of week column
        df["Day of Week"] = df["Date Opened"].dt.day_name()

        # Sort by 'Date Opened' and 'Time Opened'
        df.sort_values(["Date Opened", "Time Opened"], inplace=True)

        # Determine start and end dates
        start_date = df["Date Opened"].min().date()
        end_date = df["Date Opened"].max().date()

        # Add EntryTime Column for backtesting purposes
        df["EntryTime"] = pd.to_datetime(
            df["Date Opened"].astype(str) + " " + df["Time Opened"]
        )
        # Add column for Option Right 'C' or 'P'
        df["OptionType"] = df["Legs"].apply(
            lambda x: x.split("|")[0].strip().split(" ")[4]
        )

    else:  # BYOB BT data
        # Convert 'EntryTime' to datetime format
        df["EntryTime"] = pd.to_datetime(df["EntryTime"])

        # Add Day of week column
        df["Day of Week"] = df["EntryTime"].dt.day_name()

        # Create a 'Time' column
        df["Time"] = df["EntryTime"].dt.strftime("%H:%M:%S")

        # Sort by 'EntryTime'
        df.sort_values(["EntryTime"], inplace=True)

        # Determine start and end dates
        start_date = df["EntryTime"].min().date()
        end_date = df["EntryTime"].max().date()

    return (
        df[~df["Day of Week"].isin(weekday_exclusions)],
        start_date,
        end_date,
    )


def resize_image(image_path, size):
    """Resize the image to the specified size."""
    img = Image.open(image_path)
    img = img.resize(size, Image.LANCZOS)
    bio = BytesIO()
    img.save(bio, format="PNG")
    return bio.getvalue()


@with_gc
def run_analysis_threaded(
    files_list,
    strategy_settings,
    open_files,
):
    # initialize df_dicts
    df_dicts = {"Put/Call Comb": {}, "Puts": {}, "Calls": {}}
    for key in df_dicts:
        df_dicts[key] = {
            "All": {},
            "Mon": {},
            "Tue": {},
            "Wed": {},
            "Thu": {},
            "Fri": {},
        }

    for file in files_list:
        # check for cancel flag to stop thread
        if cancel_flag.is_set():
            cancel_flag.clear()
            results_queue.put(("-BACKTEST_CANCELED-", ""))
            return

        strategy = (
            "-SINGLE_MODE-"
            if "-SINGLE_MODE-" in strategy_settings
            else os.path.basename(file)
        )
        settings = strategy_settings[strategy]
        result_dicts = create_excel_file(file, settings, open_files)

        source = os.path.splitext(os.path.basename(file))[0]
        for right_type, day_dict in result_dicts.items():
            for day, df_dict in day_dict.items():
                df_dicts[right_type][day][source] = df_dict

    df_dicts["Best P/C"] = {
        "All": {},
        "Mon": {},
        "Tue": {},
        "Wed": {},
        "Thu": {},
        "Fri": {},
    }

    # combine the put and call dfs into 1 dict for dertmining the best time
    # from among both individual datasets
    for _day, _day_dict in df_dicts["Puts"].items():
        for _source, _df_dict in _day_dict.items():
            df_dicts["Best P/C"][_day][f"Put||{_source}"] = _df_dict

    for _day, _day_dict in df_dicts["Calls"].items():
        for _source, _df_dict in _day_dict.items():
            df_dicts["Best P/C"][_day][f"Call||{_source}"] = _df_dict

    results_queue.put(("-RUN_ANALYSIS_END-", df_dicts))
    return df_dicts


def update_strategy_settings(values, settings):
    settings.update(
        {
            "-AVG_PERIOD_1-": values["-AVG_PERIOD_1-"],
            "-PERIOD_1_WEIGHT-": values["-PERIOD_1_WEIGHT-"],
            "-AVG_PERIOD_2-": values["-AVG_PERIOD_2-"],
            "-PERIOD_2_WEIGHT-": values["-PERIOD_2_WEIGHT-"],
            "-TOP_X-": values["-TOP_X-"],
            "-CALC_TYPE-": values["-CALC_TYPE-"],
            "-AGG_TYPE-": values["-AGG_TYPE-"],
            "-MIN_TRANCHES-": values["-MIN_TRANCHES-"],
            "-MAX_TRANCHES-": values["-MAX_TRANCHES-"],
            "-BP_PER-": values["-BP_PER-"],
        }
    )

    # Initialize option settings if they don't exist
    for option in [
        "-WEEKDAY_EXCLUSIONS-",
        "-NEWS_EXCLUSIONS-",
        "-PUT_OR_CALL-",
        "-IDV_WEEKDAY-",
    ]:
        if option not in settings:
            settings[option] = [] if option.endswith("EXCLUSIONS-") else False


def validate_strategy_settings(strategy_settings):
    for strategy in strategy_settings:
        try:
            period1 = int(strategy_settings[strategy]["-AVG_PERIOD_1-"])
            period2 = int(strategy_settings[strategy]["-AVG_PERIOD_2-"])
            weight1 = float(strategy_settings[strategy]["-PERIOD_1_WEIGHT-"])
            weight2 = float(strategy_settings[strategy]["-PERIOD_2_WEIGHT-"])
            strategy_settings[strategy]["-AVG_PERIOD_1-"] = period1
            strategy_settings[strategy]["-AVG_PERIOD_2-"] = period2
            strategy_settings[strategy]["-PERIOD_1_WEIGHT-"] = weight1
            strategy_settings[strategy]["-PERIOD_2_WEIGHT-"] = weight2

            strategy_settings[strategy]["-TOP_X-"] = int(
                strategy_settings[strategy]["-TOP_X-"]
            )
            strategy_settings[strategy]["-MIN_TRANCHES-"] = int(
                strategy_settings[strategy]["-MIN_TRANCHES-"]
            )
            strategy_settings[strategy]["-MAX_TRANCHES-"] = int(
                strategy_settings[strategy]["-MAX_TRANCHES-"]
            )
            strategy_settings[strategy]["-BP_PER-"] = float(
                strategy_settings[strategy]["-BP_PER-"]
            )
        except ValueError:
            return (
                "Problem with values entered.\nPlease enter only positive whole numbers"
            )
        if period1 < 1 or period2 < 1 or period1 > period2:
            return "Please make sure both averaging periods are > 0\nand that Trailing Avg 2 is >= to Trailing Avg 1"
        if weight1 + weight2 != 100:
            return "Trailing Avg Weights should add up to 100"

    return True


@with_gc
def walk_forward_test(
    df_dicts: dict,
    path: str,
    strategy_settings: dict,
    start: dt.datetime.date = None,
    end: dt.datetime.date = None,
    initial_value: float = 100_000,
    use_scaling=False,
    export_trades=False,
):
    portfolio_mode = "-SINGLE_MODE-" not in strategy_settings
    start_date = dt.date.min
    end_date = dt.date.max
    # loop through all the source dfs
    for df_dict in df_dicts["Put/Call Comb"]["All"].values():
        # find the latest start date
        _start_date = df_dict["org_df"]["EntryTime"].min().date()
        if _start_date > start_date:
            start_date = _start_date
        # find the earliest end date
        _end_date = df_dict["org_df"]["EntryTime"].max().date()
        if _end_date < end_date:
            end_date = _end_date

    if not start:
        # find the first date the data is fully warmed up
        max_long_avg_period = max(
            [
                max(settings["-AVG_PERIOD_1-"], settings["-AVG_PERIOD_2-"])
                for settings in strategy_settings.values()
            ]
        )
        date_adv = start_date + relativedelta(months=max_long_avg_period)
        start = dt.date(date_adv.year, date_adv.month, 1)
    else:
        # use either the user input date or the first date in the dateset
        # whichever is later.  The data may not be warmed up, but the user
        # has overriden this.
        start = max(start_date, start)
    end = end_date if end is None else end

    if not portfolio_mode:
        settings = strategy_settings["-SINGLE_MODE-"]
        strats = ["All-P_C_Comb"]
        if settings["-PUT_OR_CALL-"] and settings["-IDV_WEEKDAY-"]:
            strats += ["Weekday-P_C_Comb", "All-Best_P_or_C", "Weekday-Best_P_or_C"]
        elif settings["-IDV_WEEKDAY-"]:
            strats.append("Weekday-P_C_Comb")
        elif settings["-PUT_OR_CALL-"]:
            strats.append("All-Best_P_or_C")
    else:
        strats = ["Portfolio"] + list(strategy_settings.keys())

    portfolio_metrics = {}
    for _strat in strats:
        portfolio_metrics[_strat] = {
            "Current Value": initial_value,
            "Highest Value": initial_value,
            "Max DD": 0.0,
            "Current DD": 0.0,
            "DD Days": 0,
            "Tranche Qtys": [],
            "Port Tranche Qtys": [],
            "Num Tranches": 1,
            "Port Num Tranches": 1,
            "trade log": pd.DataFrame(),
            "Win Streak": 0,
            "Loss Streak": 0,
        }

    if portfolio_mode:
        port_dict = portfolio_metrics["Portfolio"]

    # init results
    results = {}
    for strategy in portfolio_metrics:
        results[strategy] = pd.DataFrame()

    # convert weekdays from full day name to short name. i.e. Monday to Mon
    day_list = [_day[:3] for _day in weekday_list]

    current_date = start
    skip_day = False
    while current_date <= end:
        # check for cancel flag to stop thread
        if cancel_flag.is_set():
            cancel_flag.clear()
            results_queue.put(("-BACKTEST_CANCELED-", ""))
            return

        if portfolio_mode:
            # reset daily pnl for portfolio
            port_dict["Current Day PnL"] = 0

        current_weekday = current_date.strftime("%a")
        for strat, strat_dict in portfolio_metrics.items():
            if portfolio_mode and strat == "Portfolio":
                # we don't trade the portfolio, it is just the combination of all individual strats
                continue
            elif portfolio_mode:
                settings = strategy_settings[strat]
            else:
                settings = strategy_settings["-SINGLE_MODE-"]

            # reset daily pnl for individual strategy
            strat_dict["Current Day PnL"] = 0

            day_exlusions = [_day[:3] for _day in settings["-WEEKDAY_EXCLUSIONS-"]]
            # get list of news event dates to skip.
            news_date_exclusions = []
            for release, date_list in news_events.items():
                if release in settings["-NEWS_EXCLUSIONS-"]:
                    news_date_exclusions += date_list

            skip_day = False
            if (
                current_weekday in day_exlusions
                or current_weekday not in day_list
                or current_date in news_date_exclusions
            ):
                skip_day = True

            if not skip_day:
                if use_scaling:

                    def determine_num_tranches(
                        min_tranches, max_tranches, num_contracts
                    ):
                        tranches = max_tranches
                        while True:
                            if num_contracts > tranches:
                                max_tranche_qty = int(num_contracts / tranches)
                                remain_qty = num_contracts - (
                                    tranches * max_tranche_qty
                                )
                                if remain_qty >= min_tranches or remain_qty == 0:
                                    # we're done we can stay at this number of tranches with
                                    # the remainder filling up another set of at least min tranches
                                    return tranches
                                else:
                                    # we need to take a tranche away so we can try to fill up at
                                    # least 1 full set at min amount
                                    if tranches - 1 < min_tranches:
                                        # we can't reduce any further, got with what we have
                                        # even if that means we will be adding contracts below the min
                                        return tranches
                                    else:
                                        tranches -= 1
                            else:
                                return num_contracts

                    def determine_tranche_qtys(tranches):
                        tranche_qtys = []
                        for x in range(tranches):
                            if x < num_contracts % tranches:
                                # this is where we add the remaining contracts after filling up all tranches
                                tranche_qtys.append(int(num_contracts / tranches) + 1)
                            else:
                                tranche_qtys.append(int(num_contracts / tranches))
                        return tranche_qtys

                    min_tranches = settings["-MIN_TRANCHES-"]
                    max_tranches = settings["-MAX_TRANCHES-"]
                    bp_per_contract = settings["-BP_PER-"]
                    num_contracts = int(strat_dict["Current Value"] / bp_per_contract)
                    tranches = determine_num_tranches(
                        min_tranches, max_tranches, num_contracts
                    )
                    strat_dict["Num Tranches"] = tranches
                    strat_dict["Tranche Qtys"] = determine_tranche_qtys(tranches)
                    if portfolio_mode:
                        equal_value = port_dict["Current Value"] / (
                            len(portfolio_metrics) - 1
                        )
                        num_contracts = int(equal_value / bp_per_contract)
                        tranches = determine_num_tranches(
                            min_tranches, max_tranches, num_contracts
                        )
                        strat_dict["Port Num Tranches"] = tranches
                        strat_dict["Port Tranche Qtys"] = determine_tranche_qtys(
                            tranches
                        )
                else:
                    # not scaling
                    num_contracts = settings["-TOP_X-"]
                    strat_dict["Num Tranches"] = num_contracts
                    strat_dict["Tranche Qtys"] = [1 for x in range(num_contracts)]
                    strat_dict["Port Num Tranches"] = num_contracts
                    strat_dict["Port Tranche Qtys"] = [1 for x in range(num_contracts)]

                if settings["-AGG_TYPE-"] == "Monthly":
                    # date for best times should be the month prior as we don't know the future yet
                    best_time_date = current_date - relativedelta(months=1)
                else:
                    # grab from last week
                    best_time_date = current_date - relativedelta(weeks=1)

                def log_pnl_and_trades(strat_dict, num_tranches, tranche_qtys):
                    if portfolio_mode:
                        if settings["-PUT_OR_CALL-"] and settings["-IDV_WEEKDAY-"]:
                            df_dict = df_dicts["Best P/C"][current_weekday]
                        elif settings["-PUT_OR_CALL-"]:
                            df_dict = df_dicts["Best P/C"]["All"]
                        elif settings["-IDV_WEEKDAY-"]:
                            df_dict = df_dicts["Put/Call Comb"][current_weekday]
                        else:
                            df_dict = df_dicts["Put/Call Comb"]["All"]
                    else:
                        if strat == "All-P_C_Comb":
                            df_dict = df_dicts["Put/Call Comb"]["All"]
                        elif strat == "Weekday-P_C_Comb":
                            df_dict = df_dicts["Put/Call Comb"][current_weekday]
                        elif strat == "All-Best_P_or_C":
                            df_dict = df_dicts["Best P/C"]["All"]
                        elif strat == "Weekday-Best_P_or_C":
                            df_dict = df_dicts["Best P/C"][current_weekday]

                    best_times_df = get_top_times(
                        df_dict, strategy_settings, best_time_date, num_tranches
                    )

                    if portfolio_mode:
                        # filter out other sources since all sources are included
                        source = os.path.splitext(strat)[0]
                        best_times_df = best_times_df[best_times_df["Source"] == source]

                    best_times = best_times_df["Top Times"].to_list()
                    for time in best_times:
                        # get the qty for this tranche time
                        qty = tranche_qtys[best_times.index(time)]

                        full_dt = dt.datetime.combine(
                            current_date, dt.datetime.strptime(time, "%H:%M:%S").time()
                        )
                        source = best_times_df.loc[
                            best_times_df["Top Times"] == time, "Source"
                        ].values[0]

                        source_df = df_dict[source]["org_df"]
                        filtered_rows = source_df[
                            source_df["EntryTime"] == full_dt
                        ].copy()

                        if filtered_rows.empty:
                            continue

                        filtered_rows["qty"] = qty
                        filtered_rows["source"] = source

                        if is_BYOB_data(source_df):
                            gross_pnl = (
                                filtered_rows["ProfitLossAfterSlippage"].sum()
                                * 100
                                * qty
                            )
                            commissions = filtered_rows["CommissionFees"].sum() * qty
                            pnl = gross_pnl - commissions
                        else:
                            pnl = filtered_rows["P/L"].sum() * qty

                        strat_dict["Current Value"] += pnl
                        strat_dict["Current Day PnL"] += pnl
                        # log trade
                        strat_dict["trade log"] = pd.concat(
                            [strat_dict["trade log"], filtered_rows], ignore_index=True
                        )

                num_tranches = strat_dict["Num Tranches"]
                tranche_qtys = strat_dict["Tranche Qtys"]
                log_pnl_and_trades(strat_dict, num_tranches, tranche_qtys)
                if portfolio_mode:
                    num_tranches = strat_dict["Port Num Tranches"]
                    tranche_qtys = strat_dict["Port Tranche Qtys"]
                    log_pnl_and_trades(port_dict, num_tranches, tranche_qtys)

                def calc_metrics(strat_dict: dict, strat: str, results: dict) -> None:
                    # calc metrics and log the results for the day
                    if strat_dict["Current Value"] >= strat_dict["Highest Value"]:
                        strat_dict["Highest Value"] = strat_dict["Current Value"]
                        strat_dict["DD Days"] = 0
                    else:
                        # we are in Drawdown
                        dd = (
                            strat_dict["Highest Value"] - strat_dict["Current Value"]
                        ) / strat_dict["Highest Value"]
                        strat_dict["Current DD"] = dd
                        if dd > strat_dict["Max DD"]:
                            strat_dict["Max DD"] = dd
                        strat_dict["DD Days"] += 1

                    if strat_dict["Current Day PnL"] > 0:
                        strat_dict["Win Streak"] += 1
                        strat_dict["Loss Streak"] = 0
                    elif strat_dict["Current Day PnL"] < 0:
                        # tie does not change any streak
                        strat_dict["Win Streak"] = 0
                        strat_dict["Loss Streak"] += 1

                    new_row = pd.DataFrame(
                        [
                            {
                                "Date": current_date,
                                "Current Value": strat_dict["Current Value"],
                                "Highest Value": strat_dict["Highest Value"],
                                "Max DD": strat_dict["Max DD"],
                                "Current DD": strat_dict["Current DD"],
                                "DD Days": strat_dict["DD Days"],
                                "Day PnL": strat_dict["Current Day PnL"],
                                "Win Streak": strat_dict["Win Streak"],
                                "Loss Streak": strat_dict["Loss Streak"],
                                "Initial Value": initial_value,
                                "Weekday": current_weekday,
                            }
                        ]
                    )
                    results[strat] = pd.concat(
                        [results[strat], new_row], ignore_index=True
                    )

                calc_metrics(strat_dict, strat, results)
                if portfolio_mode:
                    calc_metrics(port_dict, "Portfolio", results)
            else:
                # this is a skip day just increment the DD days if needed
                if strat_dict["DD Days"] > 0:
                    strat_dict["DD Days"] += 1
                if portfolio_mode and port_dict["DD Days"] > 0:
                    port_dict["DD Days"] += 1

        current_date += dt.timedelta(1)

    for strat in portfolio_metrics:
        results[strat]["Date"] = pd.to_datetime(results[strat]["Date"])
        if export_trades:
            base_filename = f"{strat} - TradeLog_{str(uuid.uuid4())[:8]}"
            ext = ".csv"
            export_filename = get_next_filename(path, base_filename, ext)
            portfolio_metrics[strat]["trade log"].to_csv(export_filename)
    results_queue.put(("-BACKTEST_END-", results))
    return results


@with_gc
def options_window(settings) -> None:
    weekday_exclusion_checkboxes = [
        Checkbox(
            day,
            day in settings["-WEEKDAY_EXCLUSIONS-"],
            key=day,
            font=font,
            size=(6, 1),
        )
        for day in weekday_list
    ]
    news_exclusion_checkboxes = [
        Checkbox(
            release,
            release in settings["-NEWS_EXCLUSIONS-"],
            key=release,
            font=font,
            size=(11, 1),
        )
        for release in news_events
    ]
    # break into rows of 3
    news_exclusion_checkboxes = chunk_list(news_exclusion_checkboxes, 3)
    layout = [
        [
            sg.Text(
                "Economic Calendar CSV file (https://www.fxstreet.com/economic-calendar)"
            )
        ],
        [
            sg.Input(
                key="-FILE-",
                expand_x=True,
            ),
            sg.Button("Browse"),
        ],
        [
            sg.Frame(
                "Exclude Weekday",
                [weekday_exclusion_checkboxes],
                expand_x=True,
            ),
        ],
        [
            sg.Frame(
                "Exclude news",
                news_exclusion_checkboxes,
                expand_x=True,
            ),
        ],
        [
            sg.Frame(
                "Analysis Options",
                [
                    [
                        Checkbox(
                            "Put or Call",
                            settings["-PUT_OR_CALL-"],
                            key="-PUT_OR_CALL-",
                            font=font,
                            size=(6, 1),
                            tooltip="Compare selecting the best times to trade only puts or calls",
                        ),
                        Checkbox(
                            "Individual Weekday",
                            settings["-IDV_WEEKDAY-"],
                            key="-IDV_WEEKDAY-",
                            font=font,
                            size=(10, 1),
                            tooltip="Compare selecting the best times for each specific weekday to trade for that weekday",
                        ),
                    ]
                ],
                expand_x=True,
            ),
        ],
        [
            sg.Ok(),
            sg.Cancel(),
        ],
    ]
    window_size = (int(screen_size[0] * 0.4), int(screen_size[1] * 0.45))
    window = sg.Window(
        "Options",
        layout,
        no_titlebar=False,
        size=window_size,
        finalize=True,
        modal=True,
        resizable=True,
        grab_anywhere=True,
    )
    Checkbox.initial(window)
    while True:
        event, values = window.read()

        if event in (sg.WIN_CLOSED, "Cancel"):
            break

        elif event == "Browse":
            news_file = sg.popup_get_file(
                "",
                file_types=(("CSV Files", "*.csv"),),
                multiple_files=False,
                no_window=True,
            )
            window["-FILE-"].update(news_file)

        elif event == "Ok":
            settings["-WEEKDAY_EXCLUSIONS-"] = [
                day for day in weekday_list if values[day]
            ]
            settings["-NEWS_EXCLUSIONS-"] = [
                release for release in news_events if values[release]
            ]
            settings["-PUT_OR_CALL-"] = values["-PUT_OR_CALL-"]
            settings["-IDV_WEEKDAY-"] = values["-IDV_WEEKDAY-"]
            if values["-FILE-"]:
                result = import_news_events(values["-FILE-"])
                if result:
                    break
                else:
                    sg.popup_no_border(
                        "This does not appear to be a CSV from\nhttps://www.fxstreet.com/economic-calendar"
                    )
                    continue
            else:
                break

    window.close()
    Checkbox.clear_elements()


def main():

    tab_group_layout = []
    for tg in ["Put/Call Comb", "Best P/C", "Puts", "Calls"]:
        tg_layout = []
        for day in ["All", "Mon", "Tue", "Wed", "Thu", "Fri"]:
            tab = sg.Tab(
                day,
                [
                    [
                        sg.Table(
                            "",
                            ["Top Times", "Avg", "Source File"],
                            key=f"-TABLE_{day}_{tg}-",
                            expand_x=True,
                            auto_size_columns=True,
                            background_color="white",
                            alternating_row_color="lightgrey",
                            header_text_color="black",
                            header_background_color="lightblue",
                        )
                    ]
                ],
                expand_x=True,
            )
            tg_layout.append(tab)
        main_group_tab = sg.Tab(
            tg,
            [[sg.TabGroup([tg_layout], expand_x=True)]],
            expand_x=True,
        )
        tab_group_layout.append(main_group_tab)

    chart_tab = sg.Tab(
        "Charts",
        [
            [
                sg.TabGroup(
                    [
                        [
                            sg.Tab(
                                "PnL",
                                [
                                    [
                                        sg.Table(
                                            "",
                                            [
                                                "Strategy",
                                                "Final Value",
                                                "Profit",
                                                "Total Return",
                                                "CAGR",
                                                "Max DD",
                                                "Max DD Days",
                                                "MAR",
                                                "W Strk",
                                                "L Strk",
                                                "High Mo",
                                                "Low Mo",
                                            ],
                                            key="-PNL_TABLE_CHART-",
                                            expand_x=True,
                                            num_rows=4,
                                            auto_size_columns=True,
                                            background_color="lightgrey",
                                            alternating_row_color="white",
                                            header_text_color="black",
                                            header_background_color="lightblue",
                                        )
                                    ],
                                    [
                                        sg.Image(
                                            key="-PNL_CHART-",
                                            size=(
                                                int(screen_size[0] * 0.25),
                                                int(screen_size[1] * 0.25),
                                            ),
                                            expand_x=True,
                                            expand_y=True,
                                        )
                                    ],
                                ],
                            ),
                            sg.Tab(
                                "PnL by Weekday",
                                [
                                    [
                                        sg.Image(
                                            key="-WEEKDAY_PNL_CHART-",
                                            size=(
                                                int(screen_size[0] * 0.25),
                                                int(screen_size[1] * 0.25),
                                            ),
                                            expand_x=True,
                                            expand_y=True,
                                        )
                                    ],
                                ],
                            ),
                            sg.Tab(
                                "PnL by News Event",
                                [
                                    [
                                        sg.Image(
                                            key="-NEWS_PNL_CHART-",
                                            size=(
                                                int(screen_size[0] * 0.25),
                                                int(screen_size[1] * 0.25),
                                            ),
                                            expand_x=True,
                                            expand_y=True,
                                        )
                                    ],
                                ],
                            ),
                            sg.Tab(
                                "Avg PnL per News Event",
                                [
                                    [
                                        sg.Image(
                                            key="-NEWS_AVG_PNL_CHART-",
                                            size=(
                                                int(screen_size[0] * 0.25),
                                                int(screen_size[1] * 0.25),
                                            ),
                                            expand_x=True,
                                            expand_y=True,
                                        )
                                    ],
                                ],
                            ),
                        ]
                    ],
                    expand_x=True,
                    expand_y=True,
                )
            ]
        ],
    )
    tab_group_layout.append(chart_tab)

    layout = [
        [
            sg.Button("Analyze", pad=(5, 10), bind_return_key=True),
            sg.Text("  "),
            sg.pin(
                sg.ProgressBar(
                    100,
                    orientation="h",
                    size=(50, 30),
                    key="-PROGRESS-",
                    expand_x=True,
                    visible=False,
                ),
            ),
            sg.pin(sg.Button("Cancel", pad=(20, 0), visible=False)),
            sg.Push(),
            sg.Button("Options"),
            sg.Text(__version__),
        ],
        [sg.Text("Select trade log CSV file:")],
        [
            sg.Input(
                key="-FILE-",
                expand_x=True,
            ),
            sg.Button("Browse"),
        ],
        [
            Checkbox(
                "Portfolio Mode",
                False,
                key="-PORTFOLIO_MODE-",
                size=(12, 1),
                enable_events=True,
            ),
            sg.pin(
                sg.Combo(
                    [],
                    key="-STRATEGY_SELECT-",
                    readonly=True,
                    visible=False,
                    enable_events=True,
                    size=(50, 1),
                )
            ),
        ],
        [
            sg.Frame(
                "",
                [
                    [
                        sg.Text(
                            "Trailing Avg 1:",
                            tooltip="Number of months for first averaging period.\nNote: should be the shorter period",
                        ),
                        sg.Input(
                            "4",
                            key="-AVG_PERIOD_1-",
                            size=(3, 1),
                            justification="c",
                            tooltip="Number of months for first averaging period.\nNote: should be the shorter period",
                        ),
                        sg.Text("Months "),
                        sg.Text(
                            "Weight:",
                            tooltip="Weight in % for first avg period\nNote: Set to 100 for this and 0 for 2nd if only using 1 period",
                        ),
                        sg.Input(
                            "25",
                            key="-PERIOD_1_WEIGHT-",
                            size=(3, 1),
                            justification="c",
                            tooltip="Weight in % for first avg period\nNote: Set to 100 for this and 0 for 2nd if only using 1 period",
                        ),
                        sg.Text("   "),
                        sg.Text(
                            "Trailing Avg 2:",
                            tooltip="Number of months for second averaging period.\nNote: should be the longer period or same as 1",
                        ),
                        sg.Input(
                            "8",
                            key="-AVG_PERIOD_2-",
                            size=(3, 1),
                            justification="c",
                            tooltip="Number of months for second averaging period.\nNote: should be the longer period or same as 1",
                        ),
                        sg.Text("Months "),
                        sg.Text(
                            "Weight:",
                            tooltip="Weight in % for second avg period\nNote: Set to 0 to only use the 1st period",
                        ),
                        sg.Input(
                            "75",
                            key="-PERIOD_2_WEIGHT-",
                            size=(3, 1),
                            justification="c",
                            tooltip="Weight in % for second avg period\nNote: Set to 0 to only use the 1st period",
                        ),
                    ],
                    [
                        sg.Text(
                            "Select Top",
                            pad=(5, 5),
                            tooltip="Highlight the top n times for each month in the heatmap.\nWill also display the top n times below",
                        ),
                        sg.Input(
                            "5",
                            key="-TOP_X-",
                            size=(2, 1),
                            pad=(0, 0),
                            justification="c",
                            tooltip="Highlight the top n times for each month in the heatmap.\nWill also display the top n times below",
                        ),
                        sg.Text("Time Tranches", pad=(5, 0)),
                        sg.Text("   Averaging Mode"),
                        sg.Combo(
                            ["PCR", "PnL"],
                            "PCR",
                            key="-CALC_TYPE-",
                            readonly=True,
                        ),
                        sg.Text(
                            "   Aggregation Period",
                            tooltip="Aggregate the results into monthly averages or weekly\nIf doing a walkforward test the top times will be updated at this frequency.",
                        ),
                        sg.Combo(
                            ["Monthly", "Weekly"],
                            "Monthly",
                            key="-AGG_TYPE-",
                            tooltip="Aggregate the results into monthly averages or weekly\nIf doing a walkforward test the top times will be updated at this frequency.",
                            readonly=True,
                        ),
                        sg.Push(),
                        Checkbox(
                            "Open Excel files after creation",
                            False,
                            key="-OPEN_FILES-",
                            size=(20, 1),
                        ),
                    ],
                ],
                expand_x=True,
            )
        ],
        [
            sg.Frame(
                "",
                [
                    [
                        Checkbox(
                            "Perform walk-forward backtest",
                            False,
                            key="-BACKTEST-",
                            size=(19, 1),
                            tooltip="Out of sample/walk forward test.  Optimize times for prior lookback period\nand test outcome in the following month (out of sample).\nWalk forward to the next month and re-optomize times.",
                        ),
                        sg.Text(
                            "Starting Value",
                            tooltip="Porfolio Value to start from.  If using scaling the BP per contract\nwill be divided by this amount to determine the number of contracts to trade",
                        ),
                        sg.Input(
                            f"100000",
                            size=(10, 1),
                            key="-START_VALUE-",
                            justification="r",
                            tooltip="Porfolio Value to start from.  If using scaling the BP per contract\nwill be divided by this amount to determine the number of contracts to trade",
                        ),
                        sg.Text(
                            "   Start Date",
                            tooltip="Date to start test from. Leave blank to automatically\nselect the earliest available start date from the available data",
                        ),
                        sg.Input(
                            "",
                            key="-START_DATE-",
                            size=(12, 1),
                            justification="c",
                            tooltip="Date to start test from. Leave blank to automatically\nselect the earliest available start date from the available data",
                        ),
                        sg.Text(
                            " End Date",
                            tooltip="Date to end test. Leave blank to automatically\nselect the latest available end date from the available data",
                        ),
                        sg.Input(
                            "",
                            key="-END_DATE-",
                            size=(12, 1),
                            justification="c",
                            tooltip="Date to end test. Leave blank to automatically\nselect the latest available end date from the available data",
                        ),
                        sg.Push(),
                        Checkbox(
                            "Export Trades to CSV",
                            False,
                            key="-EXPORT-",
                            size=(16, 1),
                        ),
                    ],
                    [
                        Checkbox(
                            "Use Scaling",
                            False,
                            key="-SCALING-",
                            size=(10, 1),
                            tooltip="Uses scaling logic to determine the number of contracts\nto trade each day of the backtest based on current portfolio value\nand the BP per contract.",
                        ),
                        sg.Text(
                            "Min Tranches",
                            tooltip="When using scaling, this the minimum number of tranche times",
                        ),
                        sg.Input(
                            "5",
                            key="-MIN_TRANCHES-",
                            size=(3, 1),
                            justification="c",
                            tooltip="When using scaling, this the minimum number of tranche times",
                        ),
                        sg.Text(
                            "   Max Tranches",
                            tooltip="When using scaling, this the maximum number of tranche times.\nAdditonal contracts over this amount will be distributed among the available tranche times.",
                        ),
                        sg.Input(
                            "5",
                            key="-MAX_TRANCHES-",
                            size=(3, 1),
                            justification="c",
                            tooltip="When using scaling, this the maximum number of tranche times.\nAdditonal contracts over this amount will be distributed among the available tranche times.",
                        ),
                        sg.Text(
                            "   BP Per Contract",
                            tooltip="Amount of buying power to use for each contract.  This is only used to determine\nthe total number of contracts to trade each day when using scaling.",
                        ),
                        sg.Input(
                            "6000",
                            key="-BP_PER-",
                            size=(6, 1),
                            justification="r",
                            tooltip="Amount of buying power to use for each contract.  This is only used to determine\nthe total number of contracts to trade each day when using scaling.",
                        ),
                    ],
                ],
                expand_x=True,
            )
        ],
        [
            sg.TabGroup(
                [tab_group_layout],
                expand_x=True,
                key="-TAB_GROUP-",
            )
        ],
    ]
    window_size = (int(screen_size[0] * 0.6), int(screen_size[1] * 0.8))
    window = sg.Window(
        "Tranche Time Analyzer", layout, size=window_size, resizable=True, finalize=True
    )
    window["-PROGRESS-"].Widget.config(mode="indeterminate")
    Checkbox.initial(window)
    error = False
    chart_filenames = {}
    strategy_settings = {}
    test_running = False
    while True:
        event, values = window.read(timeout=100)
        if event == sg.WIN_CLOSED:
            break
        elif event == "Cancel" and test_running:
            # button will not do anything for normal analysis
            cancel_flag.set()
            window["Cancel"].update("Canceling...", disabled=True)

        elif event == "Options":
            if values["-PORTFOLIO_MODE-"]:
                selected_strategy = values["-STRATEGY_SELECT-"]
            else:
                selected_strategy = "-SINGLE_MODE-"
            options_window(strategy_settings[selected_strategy])

        elif event == "Analyze":
            files_list = values["-FILE-"].split(";")
            for file in files_list:
                file_ext = os.path.splitext(file)[1].lower()
                if file_ext != ".csv":
                    sg.popup_no_border(
                        "One or more of the selected files\ndo not appear to be a csv file!"
                    )
                    error = True
                    break
            if error:
                error = False  # reset
                continue

            if values["-PORTFOLIO_MODE-"]:
                selected_strategy = values["-STRATEGY_SELECT-"]
            else:
                selected_strategy = "-SINGLE_MODE-"

            # Save current settings for the selected strategy
            if selected_strategy not in strategy_settings:
                strategy_settings[selected_strategy] = {}
            update_strategy_settings(values, strategy_settings[selected_strategy])

            # Validate settings for all strategies
            result = validate_strategy_settings(strategy_settings)
            if type(result) == str:
                # there was an error
                sg.popup_no_border(result)
                continue

            start_date_str = values["-START_DATE-"]
            end_date_str = values["-END_DATE-"]
            if start_date_str:
                try:
                    start_date = parser.parse(start_date_str, fuzzy=True).date()
                except ValueError:
                    sg.popup_no_border(
                        "Problem parsing Start Date.\nTry entering in YYYY-MM-DD format"
                    )
                    continue
            else:
                start_date = None
            if end_date_str:
                try:
                    end_date = parser.parse(end_date_str, fuzzy=True).date()
                except ValueError:
                    sg.popup_no_border(
                        "Problem parsing End Date.\nTry entering in YYYY-MM-DD format"
                    )
                    continue
            else:
                end_date = None

            # All settings validated, proceed with analysis
            window["-PROGRESS-"].update(visible=True)
            window["Analyze"].update("Working...", disabled=True)
            window["Cancel"].update(visible=True)
            threading.Thread(
                target=lambda: run_analysis_threaded(
                    files_list,
                    strategy_settings,
                    values["-OPEN_FILES-"],
                ),
                daemon=True,
            ).start()
            test_running = True

        elif event == "Browse":
            files = sg.popup_get_file(
                "",
                file_types=(("CSV Files", "*.csv"),),
                multiple_files=True,
                no_window=True,
                files_delimiter=";",
            )
            if not files:
                # user hit cancel
                continue

            if type(files) == tuple:
                file_str = ";".join(files)
            else:
                file_str = files
            window["-FILE-"].update(file_str)

            strategy_settings.clear()  # reset strategy settings
            if values["-PORTFOLIO_MODE-"]:
                strategies = [os.path.basename(file) for file in file_str.split(";")]
                window["-STRATEGY_SELECT-"].update(values=strategies)
                if strategies:
                    # select the first strategy in the list
                    window["-STRATEGY_SELECT-"].update(value=strategies[0])
            else:
                strategies = ["-SINGLE_MODE-"]
                window["-STRATEGY_SELECT-"].update(values=[])

            # Initialize settings for each strategy
            for strategy in strategies:
                strategy_settings[strategy] = {}
                update_strategy_settings(values, strategy_settings[strategy])

            # We must continue so the GUI does not update with old values from the values dict
            continue

        elif event == "-PORTFOLIO_MODE-":
            portfolio_mode = values["-PORTFOLIO_MODE-"]
            window["-STRATEGY_SELECT-"].update(visible=portfolio_mode)

            if portfolio_mode:
                files = values["-FILE-"].split(";")
                strategies = [os.path.basename(file) for file in files]
                window["-STRATEGY_SELECT-"].update(values=strategies)
                if strategies:
                    # select the first strategy in the list
                    window["-STRATEGY_SELECT-"].update(value=strategies[0])
            else:
                strategies = ["-SINGLE_MODE-"]

            # Initialize settings for each strategy
            strategy_settings.clear()
            for strategy in strategies:
                strategy_settings[strategy] = {}
                update_strategy_settings(values, strategy_settings[strategy])

        elif event == "-STRATEGY_SELECT-":
            selected_strategy = values["-STRATEGY_SELECT-"]
            if selected_strategy in strategy_settings:
                for key, value in strategy_settings[selected_strategy].items():
                    if key in window.AllKeysDict:
                        window[key].update(format_float(value))

        elif event == "__TIMEOUT__":
            if chart_filenames:
                # Resize the image and update the element
                window_w, window_h = window.size
                image_width_max = int(window_w * 0.95)
                image_height_max = int(window_h * 0.5)
                image_width = min(
                    image_width_max, int(image_height_max / image_aspect_ratio)
                )
                image_size = (image_width, int(image_width * image_aspect_ratio))
                for chart, filename in chart_filenames.items():
                    resized_image = resize_image(filename, image_size)
                    window[chart].update(data=resized_image)

        # Update strategy settings when values change but while analysis is running
        if (
            values["-PORTFOLIO_MODE-"]
            and values["-STRATEGY_SELECT-"]
            and not test_running
        ):
            selected_strategy = values["-STRATEGY_SELECT-"]
            update_strategy_settings(values, strategy_settings[selected_strategy])
        elif not values["-PORTFOLIO_MODE-"] and not test_running:
            if "-SINGLE_MODE-" not in strategy_settings:
                strategy_settings["-SINGLE_MODE-"] = {}
            update_strategy_settings(values, strategy_settings["-SINGLE_MODE-"])

        # check if thread is done
        while True:
            try:
                result_key, results = results_queue.get(block=False)
            except queue.Empty:
                break

            if result_key == "-RUN_ANALYSIS_END-":
                df_dicts = results
                for right_type, day_dict in df_dicts.items():
                    for day, df_dict in day_dict.items():

                        top_times_df = get_top_times(df_dict, strategy_settings)
                        table_data = top_times_df.values.tolist()
                        window[f"-TABLE_{day}_{right_type}-"].update(
                            values=table_data, num_rows=len(table_data)
                        )

                if values["-BACKTEST-"]:
                    path = os.path.join(
                        os.path.dirname(files_list[0]), "data", "trade_logs"
                    )
                    os.makedirs(path, exist_ok=True)
                    threading.Thread(
                        target=lambda: walk_forward_test(
                            df_dicts,
                            path,
                            strategy_settings,
                            initial_value=float(values["-START_VALUE-"]),
                            start=start_date,
                            end=end_date,
                            use_scaling=values["-SCALING-"],
                            export_trades=values["-EXPORT-"],
                        ),
                        daemon=True,
                    ).start()

                else:
                    window["-PROGRESS-"].update(visible=False)
                    window["Cancel"].update(visible=False)
                    window["Analyze"].update("Analyze", disabled=False)
                    test_running = False

            elif result_key == "-BACKTEST_END-":
                charts = [
                    "-PNL_CHART-",
                    "-WEEKDAY_PNL_CHART-",
                    "-NEWS_PNL_CHART-",
                    "-NEWS_AVG_PNL_CHART-",
                ]
                # setup chart/plot filenames
                path = os.path.join(
                    os.path.dirname(files_list[0]), "data", "chart_images"
                )
                os.makedirs(path, exist_ok=True)
                ext = ".png"
                for chart in charts:
                    base_filename = f"Walkforward Test{chart}{str(uuid.uuid4())[:8]}"
                    chart_filenames[chart] = get_next_filename(path, base_filename, ext)

                table_data = get_pnl_plot(results, chart_filenames["-PNL_CHART-"])
                window["-PNL_TABLE_CHART-"].update(
                    values=table_data, num_rows=len(table_data)
                )

                get_weekday_pnl_chart(results, chart_filenames["-WEEKDAY_PNL_CHART-"])
                get_news_event_pnl_chart(results, chart_filenames["-NEWS_PNL_CHART-"])
                get_news_event_pnl_chart(
                    results, chart_filenames["-NEWS_AVG_PNL_CHART-"], False
                )

                for chart, filename in chart_filenames.items():
                    chart_image = resize_image(
                        filename,
                        (int(window.size[0] * 0.5), int(window.size[1] * 0.25)),
                    )
                    window[chart].update(data=chart_image)
                window["-TAB_GROUP-"].Widget.select(4)
                window["-PROGRESS-"].update(visible=False)
                window["Cancel"].update(visible=False)
                window["Analyze"].update("Analyze", disabled=False)

            elif result_key == "-BACKTEST_CANCELED-":
                window["-PROGRESS-"].update(visible=False)
                window["Cancel"].update("Cancel", disabled=False, visible=False)
                window["Analyze"].update("Analyze", disabled=False)
                test_running = False

        # move the progress bar
        if window["Analyze"].Disabled:
            window["-PROGRESS-"].Widget["value"] += 10
        else:
            window["-PROGRESS-"].Widget["value"] = 0

    window.close()


if __name__ == "__main__":
    main()
