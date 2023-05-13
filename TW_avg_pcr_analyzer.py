import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
import PySimpleGUI as sg
import os
import subprocess
import platform

icon = b"iVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAYAAABccqhmAAAQIElEQVR4Xu2dX2hcdRbHz81MpkmbSZtp122aamttYesW/yCL4CJ2QbesLyJ92CdBkcr6B0EfFnGh4oOUgu3LIoIgiyI+6SI+lIKFRVYLXSnIUrRaLWxqV23axDQ17WRmcpd01zgz2Zj2l3PP72TuJ6/NPd/z+5yTD7fk5k6SpmkqfEEAArkkkCCAXM6dQ0PgMgEEwCJAIMcEEECOh8/RIYAA2AEI5JgAAsjx8Dk6BBAAOwCBHBNAADkePkeHAAJgByCQYwIIIMfD5+gQQADsAARyTAAB5Hj4HB0CCIAdgECOCSCAHA+fo0MAAbADEMgxAQSQ4+FzdAggAHYAAjkmgAByPHyODgEEwA5AIMcEEECOh8/RIYAA2AEI5JgAAsjx8Dk6BBAAOwCBHBNAADkePkeHAAJgByCQYwIIIMfD5+gQQADsAARyTAAB5Hj4HB0CwQL428gRuTQ9BUEIQCAiget6B+WX/ZuDOwgWwO5P/yzjtYngYC6EAAQWT+COyq3y+/W/Cy6EAILRcSEE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYAAURDTzAE4hNAAPFnQAcQiEYgmgD+8q+/yoX6ZLSDEwwBCIhs698iv/nZ7cEogl8JFpzIhRCAgBsCCMDNKGgEAvYEEIA9cxIh4IYAAnAzChqBgD0BBGDPnEQIuCGAANyMgkYgYE8AAdgzJxECbgggADejoBEI2BNAAPbMSYSAGwLBAnjq9CEZa1xycxCNRv7081/LlmUDc0qlskdEGhoRbmok8qiIrJ7tp/rhITm/9xk3/Wk0Utxyowzse72l1InqqLzw7WGN8m5qVAo9sn/o7qB+EEATNgSAAIJ+iiJfhACUBoAAEIDSKpmWQQBKuBEAAlBaJdMyCEAJNwJAAEqrZFoGASjhRgAIQGmVTMsgACXcCAABKK2SaRkEoIQbASAApVUyLYMAlHAjAASgtEqmZRCAEm4EgACUVsm0TBQBvHT2qEw0pkwPmnXYA5VtMtRdnhOTypsd+CTgfSLSP3vW2rGj8v2br2SN2LR+Yf0GKT/2bEvmV7UJeWP0mGkfWYeVCyV5fM1tQTHBTwIGpXERBCDgigACcDUOmoGALQEEYMubNAi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEggWwBfVManLtG23GadtKK2U3qQ4J+Wz6qikkmacblt+U2mVlJJCU+jMx7yN2DaReVpJRAbbUqoi8k3mybYBM3NcHxQZLIA8vRFo1/ABqXWY7Pas2y6Dxb6mpflUUnk7aIn8XjQoiTzc1t4pSeU1vy0HddYviTwZdCUCaMI235OACCBotxxchAAWGgICQAD/I8AdwEI/LH7/nTsAldlwB8B/AVQWybwIAlBBjgAQgMoimRdBACrIEQACUFkk8yIIQAU5AkAAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5EQSgghwBIACVRTIvEkEABydOysXpuvlRswy8q+9aqRR650S8O35CGh32KPA95Y3S1zXzqOwPX2cllU+yxGteO5GZtzvd2pZ7XlL52LyXLAMTWSYitwdFBD8IFJTGRRCAgCsCCMDVOGgGArYEEIAtb9Ig4IoAAnA1DpqBgC0BBGDLmzQIuCKAAFyNg2YgYEsgWADnHrpXps+dse0247RVe1+V7q03z0nJw/sAPpr8WmY+7KWTvq4vrZTn1t7ZdiReCNIMBAE00UAACGBpCjDCg0DcASzNVfmh6/ZXgnEHsJTniQBUpscdAHcAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5EQSgghwBIACVRTIvggBUkCMABKCySOZFEIAKcgSAAFQWybwIAlBBjgAQgMoimReJIIDze5+R6fEx86NmGVj+wx+lcN0NcyJePHNE6mlnfQ7iI2tuaXn5yfHqOXnnu8+zxGtee233CnmwclNb7oikctC8l2wDl0siO4Migp8EDErjIghAwBUBBOBqHDQDAVsCCMCWN2kQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCWAAGx5kwYBVwSCBTDz9piJxpSrwyy2mQcq22Soe+bDJFq/xnc/IWm9ttjyrq7vf/p56VqzdrYnngNwNZ6rbCbCcwBPnT4kY41LV9mo72+f76PBRnbeIVLrLNlVXn5LCkMbZwfCC0F87+ZPdxfhSUAEsJQXRgQBLO35tXaPAFSmyR0AfwugskjmRRCACnIEgABUFsm8CAJQQY4AEIDKIpkXQQAqyBEAAlBZJPMiCEAFOQJAACqLZF4EAaggRwAIQGWRzIsgABXkCAABqCySeREEoIIcASAAlUUyLxJBAOZnJBACEFAnEPy3AOqdUBACEDAngADMkRMIAT8EEICfWdAJBMwJIABz5ARCwA8BBOBnFnQCAXMCCMAcOYEQ8EMgWAB5eh/AruEDUpPO+mSgPeu2y2Cxb3YTqx8ekplPe+qkr+KWG2Vg3+stRzpRHZUXvj3cSceUSqFH9g/dHXQmBNCEbb4HgRBA0G5FvwgBLDwCBIAALhPgDmDhHxav38EdgNJkuAPgvwBKq2RaBgEo4UYACEBplUzLIAAl3AgAASitkmkZBKCEGwEgAKVVMi2DAJRwIwAEoLRKpmUQgBJuBIAAlFbJtAwCUMKNABCA0iqZlokigDx9NuCLZ45IPe2sJwEfWXOLVAq9s4taO3ZUvn/zFdPFzTqssH6DlB97tiXmq9qEvDF6LOto0/rlQkkeX3NbUGbwg0BBaVwEAQi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEkAAtrxJg4ArAsEC+KI6JvUO+xv5DaWV0psU/8+AhkUkdTW4xTezTkS6Z8uk42NSH/5y8WUdVUh6V0hx89aWjibTugxPjTvqcvGtFKVLNi8bCCoULIA8vRAklT0i0ggC7PWiRB4VkdWz7fHnwF4ntXBfUZ4DQAALD8bzdyAAz9O5ut4QwNXxmve753sSkDsAJcDGZXgj0MLA+S9AEyMEwKPAC//I+PsO7gCUZoIAEIDSKpmWQQBKuBEAAlBaJdMyCEAJNwJAAEqrZFoGASjhRgAIQGmVTMsgACXcCAABKK2SaRkEoIQbASAApVUyLRNFAAcnTsrF6brpQbMOu6vv2paXZPyY93dJO+xR4ER+JSI/vhCkMXxSLn3wXtaITesXVl8jPTvub8kcbVyU9y+cMu0j67DlXUXZUd4UFBP8HEBQGhdBAAKuCCAAV+OgGQjYEkAAtrxJg4ArAgjA1ThoBgK2BBCALW/SIOCKAAJwNQ6agYAtgWAB5OnXgO+On5BGh/0a8J7yRunrKjVt21lJ5RPb7cs4LZGyiNzalnJeUvk442Tb8oksE5Hbg0KDBZCnF4LsGj4gtQ57/dmeddtlsNjXtDSfSipvBy2R34sGJZGH29o7Jam85rfloM76JZEng65EAE3Y8vzRYCIIIOgnyMVFCEBlDAiAOwCVRTIvggBUkCMABKCySOZFEIAKcgSAAFQWybwIAlBBjgAQgMoimRdBACrIEQACUFkk8yIIQAU5AkAAKotkXgQBqCBHAAhAZZHMiyAAFeQIAAGoLJJ5kQgCyNOHg35WHe24NwJtKq2SUlJoWtVJERkxX91sA2cedR5si6iKyDfZxppXn5nj+qDU4CcBg9K4CAIQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCWAAGx5kwYBVwSCBVA//k9Ja1OuDrPYZoo3/EKS5c1/I//finn4LcDE9JScrk0sFqGr63uSomwsreS3AD8xlWABnHvoXpk+d8bVwBfbzKq9r0r31pvnlMnDC0E+mvxaXjp7dLEIXV1/fWmlPLf2zraeeCFIMxAE0EQDASAAVwa74mYiPAjEHcAVT8flN7a/Eow7AJdjusKmEMAVgvrpb+MOgDsAlUUyL4IAVJAjAASgskjmRRCACnIEgABUFsm8CAJQQY4AEIDKIpkXQQAqyBEAAlBZJPMiCEAFOQJAACqLZF4EAaggRwAIQGWRzItEEMDFd96Q6cnvzY+aZWDvb++TrjVr50Tk4bMBT9cuyD8m/50lXvPaA4Ue2d53XVsunw3YDCT4SUDzaRIIAQioE0AA6kgpCIGlQwABLJ1Z0SkE1AkgAHWkFITA0iGAAJbOrOgUAuoEEIA6UgpCYOkQQABLZ1Z0CgF1AsECeOr0IRlrXFJvKGbB+T4ZaGTnHSId9vqzystvSWFo4yxu3gcQc/MWmx3hQSAEsNihxb0eAcTlr5uOAFR4cgfAo8Aqi2ReBAGoIEcACEBlkcyLIAAV5AgAAagsknkRBKCCHAEgAJVFMi+CAFSQIwAEoLJI5kUQgApyBIAAVBbJvAgCUEGOABCAyiKZF0EAKsgRAAJQWSTzIhEEMPM5chONzvpw0Acq22SouzxnfOO7n5C0XjMfa5aB/U8/3/L2o+PVc/LOd59nGWlee233CnmwclNb7oikctC8l2wDl0siO4Migh8FDkrjIghAwBUBBOBqHDQDAVsCCMCWN2kQcEUAAbgaB81AwJYAArDlTRoEXBFAAK7GQTMQsCUQLIAvqmNSl2nbbjNO21BaKb1JcU7KZ9VRSSXNON22/KbSKiklhdnQdHxM6sNf2jaRcVrSu0KKm7e2pEymdRmeGs842bZ8Ubpk87KBoNBgAeTphSC7hg9IrcNkt2fddhks9s0uTfXDQ3J+7zNBS+T1ouKWG2Vg3+st7Z2ojsoL3x722nJQX5VCj+wfujvoWgTQhG2+JwERQNBuRb8IASw8AgSAAC4T4A5g4R8Wr9/BHYDSZLgD4L8ASqtkWgYBKOFGAAhAaZVMyyAAJdwIAAEorZJpGQSghBsBIAClVTItgwCUcCMABKC0SqZlEIASbgSAAJRWybQMAlDCjQAQgNIqmZaJIoCDEyfl4nTd9KBZh93Vd61UCr1zYt4dPyGNDnsU+J7yRunrKs2etTF8Ui598F7WiE3rF1ZfIz077m/JHG1clPcvnDLtI+uw5V1F2VHeFBQT/CBQUBoXQQACrgggAFfjoBkI2BJAALa8SYOAKwIIwNU4aAYCtgQQgC1v0iDgigACcDUOmoGALQEEYMubNAi4IoAAXI2DZiBgSwAB2PImDQKuCCAAV+OgGQjYEkAAtrxJg4ArAgjA1ThoBgK2BBCALW/SIOCKAAJwNQ6agYAtAQRgy5s0CLgigABcjYNmIGBLAAHY8iYNAq4IIABX46AZCNgSQAC2vEmDgCsCCMDVOGgGArYEEIAtb9Ig4IoAAnA1DpqBgC0BBGDLmzQIuCKAAFyNg2YgYEsAAdjyJg0CrgggAFfjoBkI2BJAALa8SYOAKwIIwNU4aAYCtgT+AxBxhcTqHAGHAAAAAElFTkSuQmCC"
version = "v.1.1.2"


def calculate_pcr(df: pd.DataFrame) -> float:
    if df.columns[0] == "Date Opened":  # OO BT data
        return df["P/L"].sum() / (df["Premium"] * df["No. of Contracts"]).sum()
    elif df.columns[0] == "TradeID":  # BYOB BT data
        df["P/L"] = df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
        return df["P/L"].sum() / df["Premium"].sum()
    else:
        raise ValueError("Unknown dataset type")


def calculate_avg_pnl(df: pd.DataFrame) -> float:
    if df.columns[0] == "Date Opened":  # OO BT data
        return df["P/L"].sum() / df["No. of Contracts"].sum()
    elif df.columns[0] == "TradeID":  # BYOB BT data
        df["P/L"] = (df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100) * 100
        return df["P/L"].mean()
    else:
        raise ValueError("Unknown dataset type")


def get_day_of_week_sheets(
    df: pd.DataFrame,
    writer: pd.ExcelWriter,
    df_labels: pd.DataFrame,
    short_avg_period,
    long_avg_period,
    short_weight,
    long_weight,
) -> list:
    worksheets = []  # place all DOW worksheets here

    # Filter the DataFrame for the current day of the week
    for day in reversed(df["Day of Week"].unique().tolist()):
        df_day = df[df["Day of Week"] == day]

        # Group by month and time
        if df.columns[0] == "Date Opened":  # OO BT data
            df_day_grouped = df_day.groupby(
                [df_day["Date Opened"].dt.to_period("M"), "Time Opened"]
            )
        elif df.columns[0] == "TradeID":  # BYOB BT data
            df_day_grouped = df_day.groupby(
                [df_day["EntryTime"].dt.to_period("M"), "Time"]
            )
        else:
            raise ValueError("Unknown dataset type")

        # Calculate PCR for each group
        df_day_pcr = df_day_grouped.apply(calculate_pcr)

        # Unstack if it's a MultiIndex
        if isinstance(df_day_pcr.index, pd.MultiIndex):
            df_day_pcr = df_day_pcr.unstack()

        # Calculate rolling averages
        df_pcr_day_short_avg = df_day_pcr.rolling(short_avg_period, min_periods=1).mean()
        df_pcr_day_long_avg = df_day_pcr.rolling(long_avg_period, min_periods=1).mean()

        # Calculate weighted average
        df_pcr_day_weighted_avg = (
            short_weight * df_pcr_day_short_avg + long_weight * df_pcr_day_long_avg
        )

        # Sort the data in descending order by date
        df_pcr_day_weighted_avg.sort_index(ascending=False, inplace=True)

        # Convert df_pcr_day_weighted_avg to a DataFrame if it's a Series
        if isinstance(df_pcr_day_weighted_avg, pd.Series):
            df_pcr_day_weighted_avg = df_pcr_day_weighted_avg.to_frame()

        # round to 3 decimal place which will later be displayed from .xxx to xx.x%
        df_pcr_day_weighted_avg = df_pcr_day_weighted_avg.applymap(lambda x: round(x, 3))

        # Concatenate the output DataFrame and the weighted PCR DataFrame
        df_day_output = pd.concat([df_labels, df_pcr_day_weighted_avg], axis=1)

        # create worksheet
        df_day_output.to_excel(writer, sheet_name=day, index=False)
        worksheets.append(writer.sheets[day])

    return worksheets


def analyze(file, short_avg_period, long_avg_period, short_weight, long_weight, top_x):
    # Load the CSV file
    try:
        df = pd.read_csv(file)
    except UnicodeDecodeError:
        sg.popup_no_border(
            "This does not appear to be a backtest results\nCSV from either OptionOmega"
            " or BYOB.\n\nPlease choose a different file"
        )
        return

    # Determine which type of data, OptionOmega or BYOB
    if df.columns[0] == "Date Opened":  # OO BT data
        # Convert 'Date Opened' to datetime format
        df["Date Opened"] = pd.to_datetime(df["Date Opened"])

        # Add Day of week column
        df["Day of Week"] = df["Date Opened"].dt.day_name()

        # Sort by 'Date Opened' and 'Time Opened'
        df.sort_values(["Date Opened", "Time Opened"], inplace=True)

        # Group by month and 'Time Opened'
        df_grouped = df.groupby([df["Date Opened"].dt.to_period("M"), "Time Opened"])

        # Determine start and end dates
        start_date = df["Date Opened"].min().date()
        end_date = df["Date Opened"].max().date()

    elif df.columns[0] == "TradeID":  # BYOB BT data
        # Convert 'EntryTime' to datetime format
        df["EntryTime"] = pd.to_datetime(df["EntryTime"])

        # Add Day of week column
        df["Day of Week"] = df["EntryTime"].dt.day_name()

        # Create a 'Time' column
        df["Time"] = df["EntryTime"].dt.strftime("%H:%M:%S")

        # Sort by 'EntryTime'
        df.sort_values(["EntryTime"], inplace=True)

        # Group by month and time
        df_grouped = df.groupby([df["EntryTime"].dt.to_period("M"), "Time"])

        # Determine start and end dates
        start_date = df["EntryTime"].min().date()
        end_date = df["EntryTime"].max().date()

    else:
        sg.popup_no_border(
            "This does not appear to be a backtest results\nCSV from either OptionOmega"
            " or BYOB.\n\nPlease choose a different file"
        )
        return

    # Calculate PCR for each group
    df_pcr = df_grouped.apply(calculate_pcr)
    df_avg_pnl = df_grouped.apply(calculate_avg_pnl)

    # Unstack if it's a MultiIndex
    if isinstance(df_pcr.index, pd.MultiIndex):
        df_pcr = df_pcr.unstack()

    # Calculate rolling averages
    df_pcr_short_avg = df_pcr.rolling(short_avg_period, min_periods=1).mean()
    df_pcr_long_avg = df_pcr.rolling(long_avg_period, min_periods=1).mean()

    # Calculate weighted average
    df_pcr_weighted_avg = short_weight * df_pcr_short_avg + long_weight * df_pcr_long_avg

    # Calculate 1-month average PCR
    df_pcr_1mo_avg = df_pcr.rolling(1, min_periods=1).mean()

    # Sort the data in descending order by date
    df_pcr_weighted_avg.sort_index(ascending=False, inplace=True)
    df_pcr_1mo_avg.sort_index(ascending=False, inplace=True)

    # Convert df_pcr_weighted_avg to a DataFrame if it's a Series
    if isinstance(df_pcr_weighted_avg, pd.Series):
        df_pcr_weighted_avg = df_pcr_weighted_avg.to_frame()

    # round to 3 decimal place which will later be displayed from .xxx to xx.x%
    df_pcr_weighted_avg = df_pcr_weighted_avg.applymap(lambda x: round(x, 3))
    df_pcr_1mo_avg = df_pcr_1mo_avg.applymap(lambda x: round(x, 3))

    # Create a new DataFrame for output, adding date range as index
    df_output_labels = pd.DataFrame(index=df_pcr_weighted_avg.index)
    for i, (date, row) in enumerate(df_pcr_weighted_avg.iterrows()):
        current_month_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
        previous_month_start = (
            current_month_end - pd.DateOffset(months=long_avg_period - 1)
        ).replace(day=1)
        if i == 0:
            date_range_label = f"{end_date} - {previous_month_start.date()}"
        elif i == len(df_pcr_weighted_avg) - 1:
            date_range_label = f"{current_month_end.date()} - {start_date}"
        else:
            date_range_label = (
                f"{current_month_end.date()} - {previous_month_start.date()}"
            )
        df_output_labels.loc[date, "Date Range"] = date_range_label

    df_output_1mo_avg = pd.DataFrame(index=df_pcr_1mo_avg.index)
    for i, (date, row) in enumerate(df_pcr_1mo_avg.iterrows()):
        current_month_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
        previous_month_start = current_month_end.replace(day=1)
        if i == 0:
            date_range_label = f"{end_date} - {previous_month_start.date()}"
        elif i == len(df_pcr_1mo_avg) - 1:
            date_range_label = f"{current_month_end.date()} - {start_date}"
        else:
            date_range_label = (
                f"{current_month_end.date()} - {previous_month_start.date()}"
            )
        df_output_1mo_avg.loc[date, "Date Range"] = date_range_label

    # Concatenate the output DataFrame and the weighted PCR DataFrame
    df_output = pd.concat([df_output_labels, df_pcr_weighted_avg], axis=1)

    # Concatenate the output DataFrame and the 1mo PCR DataFrame
    df_output_1mo_avg = pd.concat([df_output_1mo_avg, df_pcr_1mo_avg], axis=1)

    # path and orginal filename
    path = os.path.dirname(file)
    org_filename = os.path.basename(file)[:-4]

    # Create filename
    filename = os.path.join(
        path,
        f"{org_filename}-TWAvg_{short_avg_period}mo-{long_avg_period}mo_{start_date}-{end_date}.xlsx",
    )

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
        # Write the DataFrame to an Excel file
        df_output.to_excel(writer, sheet_name="TW_PCR", index=False)
        df_output_1mo_avg.to_excel(writer, sheet_name="1mo Avg PCR", index=False)

        # Get the xlsxwriter workbook
        workbook = writer.book

        # Get the day of the week sheets
        worksheets = get_day_of_week_sheets(
            df,
            writer,
            df_output_labels,
            short_avg_period,
            long_avg_period,
            short_weight,
            long_weight,
        )
        worksheets.append(writer.sheets["TW_PCR"])
        worksheets.append(writer.sheets["1mo Avg PCR"])

        # Set the PCR columns to percentage format
        percent_format = workbook.add_format({"num_format": "0.0%", "align": "center"})
        top_x_format = workbook.add_format({"bold": 1, "font_color": "#FFFFFF"})  # white
        for row in range(
            2, len(df_output) + 2
        ):  # +2 because Excel's index starts from 1 and there is a header row
            for worksheet in worksheets:
                # Apply a conditional format to the PCR cells in the current row
                worksheet.conditional_format(
                    f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                    {
                        "type": "3_color_scale",
                        "min_color": "red",
                        "mid_color": "yellow",
                        "max_color": "green",
                    },
                )
                # Format top x values in bold white text
                if top_x > 0:
                    worksheet.conditional_format(
                        f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                        {
                            "type": "top",
                            "value": top_x,
                            "format": top_x_format,
                        },
                    )
                worksheet.set_row(row - 1, None, percent_format)

        # Adjust the column widths
        for column in df_output:
            column_length = max(
                df_output[column].astype(str).map(len).max() + 1, len(column) + 1
            )
            col_idx = df_output.columns.get_loc(column)
            for worksheet in worksheets:
                worksheet.set_column(col_idx, col_idx, column_length)

    # open file in excel
    if platform.system() == "Windows":
        subprocess.Popen(["cmd", "/c", "start", filename], shell=True)
    elif platform.system() == "Darwin":  # This is the value returned for macOS
        subprocess.Popen(["open", filename])
    else:
        print("Unsupported platform: ", platform.system())


def main():
    sg.theme("Reddit")
    sg.SetOptions(font=("Arial", 12), icon=icon)

    layout = [
        [sg.Text("Select trade log CSV file:")],
        [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("CSV Files", "*.csv"),))],
        [
            sg.Frame(
                "",
                [
                    [
                        sg.Text("Trailing Avg 1:"),
                        sg.Input(
                            "3", key="-AVG_PERIOD_1-", size=(3, 1), justification="c"
                        ),
                        sg.Text("Weight:"),
                        sg.Input(
                            "75", key="-PERIOD_1_WEIGHT-", size=(3, 1), justification="c"
                        ),
                    ],
                    [
                        sg.Text("Trailing Avg 2:"),
                        sg.Input(
                            "10", key="-AVG_PERIOD_2-", size=(3, 1), justification="c"
                        ),
                        sg.Text("Weight:"),
                        sg.Input(
                            "25", key="-PERIOD_2_WEIGHT-", size=(3, 1), justification="c"
                        ),
                    ],
                    [
                        sg.Text("Highlight Top", pad=(5, 5)),
                        sg.Input(
                            "5", key="-TOP_X-", size=(2, 1), pad=(0, 0), justification="c"
                        ),
                        sg.Text("Values with White Text", pad=(5, 0)),
                    ],
                ],
            )
        ],
        [
            sg.Button("Analyze", pad=(5, 10)),
            sg.Button("Cancel"),
            sg.Text("", size=(30, 1)),
            sg.Text(version),
        ],
    ]

    window = sg.Window("Trailing Weighted PCR Analyzer", layout, resizable=True)

    while True:
        event, values = window.read()

        if event == "Cancel" or event == sg.WIN_CLOSED:
            break

        if event == "Analyze":
            print(os.path.splitext(values["-FILE-"]))
            if os.path.splitext(values["-FILE-"])[1].lower() != ".csv":
                sg.popup_no_border("Please select a csv file")
            else:
                analyze(
                    values["-FILE-"],
                    int(values["-AVG_PERIOD_1-"]),
                    int(values["-AVG_PERIOD_2-"]),
                    float(values["-PERIOD_1_WEIGHT-"]) / 100,
                    float(values["-PERIOD_2_WEIGHT-"]) / 100,
                    int(values["-TOP_X-"]),
                )

    window.close()


if __name__ == "__main__":
    main()
